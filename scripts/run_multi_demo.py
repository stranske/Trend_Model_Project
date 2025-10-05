#!/usr/bin/env python
"""Run the multi-period demo using the generated data.

This script exercises the Phase‑2 multi-period engine by running
``multi_period.run`` to obtain a collection of ``score_frame`` objects and then
feeding them through ``run_schedule`` with a selector and weighting scheme.

Type Hygiene Notes
------------------
This file is intentionally verbose and exercises a very broad surface area of
the code base for CI. We add lightweight type aliases and Protocols so mypy
can reason about common structures without over-specifying every intermediate
object. The goal is clarity + reduced noise, not full strict coverage.
"""

from __future__ import annotations

import copy
import importlib
import json
import os
import shutil
import subprocess
import sys
from dataclasses import fields
from pathlib import Path
from typing import Any, Mapping, Protocol, Sequence, TypedDict, cast

# Add src to path before any project imports (acceptable for lint since it's
# still above third-party imports). This keeps environment manipulation minimal.
ROOT = Path(__file__).resolve().parent.parent
SRC_PATH = ROOT / "src"
if str(SRC_PATH) not in sys.path:
    sys.path.insert(0, str(SRC_PATH))


# Opt into the guarded interpreter bootstrap unless explicitly disabled.
def setup_trend_model_env():
    os.environ.setdefault("TREND_MODEL_SITE_CUSTOMIZE", "1")


setup_trend_model_env()
# Standard library done; third-party imports
import numpy as np
import openpyxl
import pandas as pd

# Project imports (kept together; E402 no longer triggered)
import trend_analysis as ta
import yaml
from trend_analysis import (
    cli,
    export,
    gui,
    metrics,
    pipeline,
    run_analysis,
    run_multi_analysis,
)
from trend_analysis.config import Config, load
from trend_analysis.config.models import ConfigProtocol as _ConfigProto
from trend_analysis.core import rank_selection as rs
from trend_analysis.core.rank_selection import RiskStatsConfig, rank_select_funds
from trend_analysis.data import ensure_datetime, identify_risk_free_fund, load_csv
from trend_analysis.multi_period import run as run_mp
from trend_analysis.multi_period import run_schedule, scheduler
from trend_analysis.multi_period.engine import Portfolio, SelectorProtocol
from trend_analysis.multi_period.replacer import Rebalancer
from trend_analysis.selector import RankSelector, ZScoreSelector
from trend_analysis.weighting import (
    AdaptiveBayesWeighting,
    BaseWeighting,
    EqualWeight,
    ScorePropBayesian,
    ScorePropSimple,
)

FAST_SENTINEL = ROOT / "demo/.fast_demo_mode"
if FAST_SENTINEL.exists():  # pragma: no cover - short-circuit path
    try:
        FAST_SENTINEL.unlink()
    except OSError:
        pass
    (ROOT / "demo/exports").mkdir(parents=True, exist_ok=True)
    sys.exit(0)

# ---------------------------------------------------------------------------
# Type aliases / structural protocols
# ---------------------------------------------------------------------------


class _StatsLike(Protocol):  # minimal protocol matching pipeline._Stats
    cagr: float
    vol: float
    sharpe: float
    sortino: float
    information_ratio: float
    max_drawdown: float


PeriodTuple = tuple[str, str, str, str]


class ResultDict(TypedDict, total=False):
    period: PeriodTuple
    score_frame: pd.DataFrame
    selected_funds: list[str]
    out_ew_stats: _StatsLike
    out_sample_stats: dict[str, _StatsLike]
    benchmark_ir: dict[str, Mapping[str, float]]


class PortfolioConfig(TypedDict, total=False):
    selection_mode: str
    random_n: int
    manual_list: list[str]


class SampleSplitConfig(TypedDict):
    in_start: str
    in_end: str
    out_start: str
    out_end: str


ScoreFrame = pd.DataFrame  # semantic alias for clarity
StatsMap = Mapping[str, _StatsLike]
MetricSeries = pd.Series

# (widgets and metrics imported within functions where needed)


def _check_generate_demo() -> None:
    """Run generate_demo.py in both modes."""
    csv = Path("demo/demo_returns.csv")
    xlsx = csv.with_suffix(".xlsx")
    csv.unlink(missing_ok=True)
    xlsx.unlink(missing_ok=True)
    subprocess.run(
        [sys.executable, "scripts/generate_demo.py", "--no-xlsx"],
        check=True,
        shell=False,
    )
    if not csv.exists() or xlsx.exists():
        raise SystemExit("generate_demo --no-xlsx failed")
    subprocess.run(
        [sys.executable, "scripts/generate_demo.py"], check=True, shell=False
    )
    if not xlsx.exists():
        raise SystemExit("generate_demo missing Excel")


def _check_generate_demo_help() -> None:
    """Ensure the dataset generator prints usage information."""
    proc = subprocess.run(
        [sys.executable, "scripts/generate_demo.py", "--help"],
        check=True,
        capture_output=True,
        text=True,
        shell=False,
    )
    if "Generate demo return series" not in proc.stdout:
        raise SystemExit("generate_demo --help missing text")


def _check_demo_data(cfg: Config) -> pd.DataFrame:  # use concrete Config for model_dump
    """Validate the generated demo CSV and return the DataFrame."""
    df = load_csv(cfg.data["csv_path"])
    if df is None:
        raise SystemExit("Failed to load demo CSV")
    df = ensure_datetime(df)
    if not df["Date"].is_monotonic_increasing:
        raise SystemExit("Demo dataset not sorted by date")
    if df.shape != (120, 22):
        raise SystemExit("Demo dataset shape mismatch")
    if df["Date"].isnull().any():
        raise SystemExit("Demo dataset contains invalid dates")
    mgr_cols = [c for c in df.columns if c != "Date"]
    if len(mgr_cols) != 21:
        raise SystemExit("Demo dataset manager count mismatch")
    first = df["Date"].iloc[0]
    last = df["Date"].iloc[-1]
    expected_last = first + pd.offsets.MonthEnd(119)
    if last != expected_last:
        raise SystemExit("Demo dataset date range mismatch")
    xlsx_path = Path(cfg.data["csv_path"]).with_suffix(".xlsx")
    if not xlsx_path.exists():
        raise SystemExit("Demo Excel file missing")
    df_xlsx = pd.read_excel(xlsx_path)
    if df_xlsx.shape != df.shape:
        raise SystemExit("Demo Excel shape mismatch")
    return df


def _check_schedule(
    score_frames: Mapping[str, pd.DataFrame],
    selector: SelectorProtocol,
    weighting: BaseWeighting,
    cfg: Config | _ConfigProto,
    *,
    rank_column: str | None = None,
) -> Portfolio:
    # Ensure we have access to model_dump (present on concrete Config and fallback)
    cfg_dump: dict[str, Any]
    if hasattr(cfg, "model_dump"):
        cfg_dump = getattr(cfg, "model_dump")()  # type: ignore[no-any-return]
    else:  # pragma: no cover - defensive
        cfg_dump = cast(dict[str, Any], getattr(cfg, "__dict__", {}))
    rebalancer = Rebalancer(cfg_dump)
    pf = run_schedule(
        score_frames,
        selector,
        weighting,
        rank_column=rank_column,
        rebalancer=rebalancer,
    )
    if len(pf.history) != len(score_frames):
        raise SystemExit("Weight schedule did not cover all periods")
    weights = list(pf.history.values())
    if len(weights) > 1 and all(w.equals(weights[0]) for w in weights[1:]):
        print("Warning: weights did not change across periods")
    if hasattr(weighting, "mean") and hasattr(weighting, "tau"):
        if getattr(weighting, "mean") is None or getattr(weighting, "tau") is None:
            raise SystemExit("Weighting state not updated")
    return pf


def _check_gui(cfg_path: str) -> None:
    """Exercise basic GUI helpers."""
    store = gui.ParamStore.from_yaml(Path(cfg_path))
    gui.save_state(store)
    loaded = gui.load_state()
    if loaded.cfg != store.cfg:
        raise SystemExit("GUI state roundtrip failed")
    gui.reset_weight_state(loaded)
    if store.to_dict() != store.cfg:
        raise SystemExit("ParamStore.to_dict mismatch")
    gui.discover_plugins()
    cfg_dict = gui.build_config_dict(store)
    cfg_obj = gui.build_config_from_store(store)
    dumped = cfg_obj.model_dump()
    for k, v in cfg_dict.items():
        if dumped.get(k) != v:
            raise SystemExit("build_config_from_store mismatch")

    class _DummyPlugin:
        pass

    gui.register_plugin(_DummyPlugin)
    if _DummyPlugin not in list(gui.iter_plugins()):
        raise SystemExit("Plugin registration failed")
    builtin_cfgs = gui.list_builtin_cfgs()
    if "demo" not in builtin_cfgs:
        raise SystemExit("list_builtin_cfgs missing demo.yml")
    if "defaults" not in builtin_cfgs:
        raise SystemExit("list_builtin_cfgs missing defaults.yml")
    if builtin_cfgs != sorted(builtin_cfgs):
        raise SystemExit("list_builtin_cfgs not sorted")
    import ipywidgets as widgets

    from trend_analysis.core.rank_selection import build_ui
    from trend_analysis.gui import app as gui_app

    ui = build_ui()
    if not isinstance(ui, widgets.Widget):
        raise SystemExit("build_ui did not return a Widget")

    # exercise more GUI construction helpers
    man = gui_app._build_manual_override(store)
    weight = gui_app._build_weighting_options(store)
    step0 = gui_app._build_step0(store)
    rank = gui_app._build_rank_options(store)
    for widget in (man, weight, step0, rank):
        if not isinstance(widget, widgets.Widget):
            raise SystemExit("GUI builder did not return a Widget")

    # exercise weight state persistence
    store.weight_state = {"dummy": [1, 2, 3]}
    gui.save_state(store)
    loaded_ws = gui.load_state()
    if loaded_ws.weight_state != store.weight_state:
        raise SystemExit("weight_state roundtrip failed")

    # verify launch() returns a Widget
    app = gui.launch()
    if not isinstance(app, widgets.Widget):
        raise SystemExit("launch() did not return a Widget")


def _check_datagrid_override() -> None:
    """Exercise DataGrid path in manual override builder."""

    import sys
    import types

    import ipywidgets as widgets

    from trend_analysis.gui import app as gui_app

    mod = types.ModuleType("ipydatagrid")

    class DummyGrid(widgets.Widget):  # type: ignore[misc]
        def __init__(self, *a: object, **k: object) -> None:
            super().__init__()

    mod.DataGrid = DummyGrid  # type: ignore[attr-defined]
    sys.modules["ipydatagrid"] = mod
    try:
        store = gui.ParamStore()
        widget = gui_app._build_manual_override(store)
        if not isinstance(widget, widgets.Widget):
            raise SystemExit("DataGrid override path failed")
    finally:
        sys.modules.pop("ipydatagrid", None)


def _check_plugin_discovery() -> None:
    """Validate discover_plugins registers entry points."""

    import importlib.metadata as md
    import sys
    import types

    from trend_analysis.gui import plugins

    dummy_mod = types.ModuleType("demo_plugin")

    class DemoPlugin:
        pass

    dummy_mod.DemoPlugin = DemoPlugin  # type: ignore[attr-defined]
    sys.modules["demo_plugin"] = dummy_mod
    ep = md.EntryPoint("demo", "demo_plugin:DemoPlugin", "trend_analysis.gui_plugins")
    orig = md.entry_points

    def fake(group: str | None = None) -> list[md.EntryPoint]:
        if group == "trend_analysis.gui_plugins":
            return [ep]
        return []

    md.entry_points = fake  # type: ignore[assignment]
    try:
        plugins._PLUGIN_REGISTRY.clear()
        gui.discover_plugins()
        if DemoPlugin not in list(gui.iter_plugins()):
            raise SystemExit("Plugin discovery failed")
    finally:
        md.entry_points = orig  # type: ignore[assignment]
        sys.modules.pop("demo_plugin", None)
        plugins._PLUGIN_REGISTRY.clear()


def _check_selection_modes(cfg: Config) -> None:
    """Verify legacy selection modes still operate."""
    base = cfg.model_dump()
    for mode in ("all", "random", "manual"):
        cfg_copy = Config(**base)
        cfg_copy.portfolio["selection_mode"] = mode
        if mode == "random":
            cfg_copy.portfolio["random_n"] = 4
        if mode == "manual":
            cfg_copy.portfolio["manual_list"] = ["Mgr_01", "Mgr_02"]
        res = pipeline.run_full(cfg_copy)
        sel = res.get("selected_funds")
        if not sel:
            raise SystemExit(f"{mode} mode produced no funds")


def _check_cli_env(cfg_path: str) -> None:
    """Invoke the CLI using the TREND_CFG environment variable."""
    env = os.environ.copy()
    env["TREND_CFG"] = cfg_path
    env["PYTHONPATH"] = f"{ROOT / 'src'}:{env.get('PYTHONPATH', '')}"
    env.setdefault("TREND_MODEL_SITE_CUSTOMIZE", "1")
    subprocess.run(
        [sys.executable, "-m", "trend_analysis.run_analysis", "--detailed"],
        check=True,
        env=env,
        shell=False,
    )
    os.environ["TREND_CFG"] = cfg_path
    rc = run_analysis.main([])
    os.environ.pop("TREND_CFG", None)
    if rc != 0:
        raise SystemExit("run_analysis.main env failed")


def _check_cli_env_multi(cfg_path: str) -> None:
    """Invoke the multi-period CLI using the TREND_CFG variable."""
    env = os.environ.copy()
    env["TREND_CFG"] = cfg_path
    env["PYTHONPATH"] = f"{ROOT / 'src'}:{env.get('PYTHONPATH', '')}"
    env.setdefault("TREND_MODEL_SITE_CUSTOMIZE", "1")
    subprocess.run(
        [sys.executable, "-m", "trend_analysis.run_multi_analysis", "--detailed"],
        check=True,
        env=env,
        shell=False,
    )
    os.environ["TREND_CFG"] = cfg_path
    rc = run_multi_analysis.main([])
    os.environ.pop("TREND_CFG", None)
    if rc != 0:
        raise SystemExit("run_multi_analysis.main env failed")


def _check_cli(cfg_path: str, csv_path: str) -> None:
    """Exercise the simple CLI wrapper."""
    rc = cli.main(["--check"])
    if rc != 0:
        raise SystemExit("CLI --check failed")
    rc = cli.main(["run", "-c", cfg_path, "-i", csv_path])
    if rc != 0:
        raise SystemExit("CLI run failed")


def _check_misc(
    cfg_path: str,
    cfg: Config,
    results: Sequence[dict[str, Any]],
) -> None:
    """Exercise smaller utility modules."""
    import asyncio

    from trend_analysis import metrics

    if "annual_return" not in metrics.available_metrics():
        raise SystemExit("Metrics registry incomplete")

    all_metrics = rs.canonical_metric_list()
    if "Sharpe" not in all_metrics or "AnnualReturn" not in all_metrics:
        raise SystemExit("canonical_metric_list default list incorrect")

    # scheduler.generate_periods should agree with the results length
    periods = scheduler.generate_periods(cfg.model_dump())
    if len(periods) != len(results):
        raise SystemExit("Scheduler period count mismatch")

    store = gui.ParamStore.from_yaml(Path(cfg_path))
    if gui.build_config_from_store(store).version != cfg.version:
        raise SystemExit("Config build roundtrip failed")

    called: list[int] = []

    @gui.debounce(50)  # type: ignore[misc]
    async def ping(val: int) -> None:
        called.append(val)

    async def drive() -> None:
        await ping(1)
        await ping(2)
        await asyncio.sleep(0.1)

    asyncio.run(drive())
    if called != [2]:
        raise SystemExit("debounce failed")

    os.environ["TREND_CFG"] = cfg_path
    cfg_env = load(None)
    os.environ.pop("TREND_CFG", None)
    if cfg_env.version != cfg.version:
        raise SystemExit("TREND_CFG not honoured")
    if cfg_env.jobs != 1 or cfg_env.checkpoint_dir != "demo/checkpoints":
        raise SystemExit("Config optional fields not parsed")

    df_demo = load_csv(cfg.data["csv_path"])
    df_demo = ensure_datetime(df_demo)
    sf_single = pipeline.single_period_run(
        df_demo[["Date", "Mgr_01", "Mgr_02"]],
        str(cfg.sample_split["in_start"]),
        str(cfg.sample_split["in_end"]),
    )
    if sf_single.attrs.get("period") != (
        str(cfg.sample_split["in_start"]),
        str(cfg.sample_split["in_end"]),
    ):
        raise SystemExit("single_period_run metadata mismatch")

    pr = pipeline.calc_portfolio_returns(
        np.array([0.5, 0.5]),
        df_demo[["Mgr_01", "Mgr_02"]].iloc[:4],
    )
    if len(pr) != 4:
        raise SystemExit("calc_portfolio_returns length mismatch")

    # dynamic metric registration
    @metrics.register_metric("dummy_metric")  # type: ignore[misc]
    def _dummy(series: pd.Series) -> float:
        return float(series.mean())

    if "dummy_metric" not in metrics.available_metrics():
        raise SystemExit("Metric registration failed")
    metrics._METRIC_REGISTRY.pop("dummy_metric", None)

    clist = rs.canonical_metric_list(["sharpe_ratio", "max_drawdown"])
    if clist != ["Sharpe", "MaxDrawdown"]:
        raise SystemExit("canonical_metric_list failed")

    # direct calls to helper functions for coverage
    series = pd.Series([1.0, 2.0, 3.0])
    ranked = rs._apply_transform(series, mode="rank")
    if ranked.iloc[0] != 3:
        raise SystemExit("_apply_transform failed")

    df_tmp = pd.DataFrame({"A": [0.01, 0.02], "B": [0.02, 0.03]})
    scores = rs._compute_metric_series(df_tmp, "AnnualReturn", RiskStatsConfig())
    if len(scores) != 2:
        raise SystemExit("_compute_metric_series failed")

    # verify pipeline alias handling
    stats_cls = getattr(pipeline, "Stats")
    if stats_cls is not pipeline._Stats:
        raise SystemExit("pipeline.Stats alias failed")


def _check_config_dump(cfg: Config) -> None:
    """Verify ``Config`` serialisation helpers."""
    data = cfg.model_dump()
    dumped = cfg.model_dump_json()
    if not isinstance(dumped, str) or "version" not in dumped:
        raise SystemExit("model_dump_json output invalid")
    cfg2 = Config(**data)
    if cfg2.version != cfg.version:
        raise SystemExit("Config model_dump roundtrip failed")


def _check_default_load() -> None:
    """Ensure ``load(None)`` falls back to ``defaults.yml``."""
    os.environ.pop("TREND_CFG", None)
    cfg_def = load(None)
    cfg_ref = load("config/defaults.yml")
    if cfg_def.version != cfg_ref.version:
        raise SystemExit("Config default load mismatch")


def _check_rebalancer_logic() -> None:
    """Verify Rebalancer triggers drop and add events."""
    reb = Rebalancer({})
    prev = pd.Series({"A": 0.5, "B": 0.5})
    sf1 = pd.DataFrame({"zscore": [0.0, -1.1]}, index=["A", "B"])
    w1 = reb.apply_triggers(prev, sf1)
    sf2 = pd.DataFrame({"zscore": [0.0, -1.2]}, index=["A", "B"])
    w2 = reb.apply_triggers(w1, sf2)
    sf3 = pd.DataFrame({"zscore": [0.0, 1.2]}, index=["A", "C"])
    w3 = reb.apply_triggers(w2, sf3)
    if "B" in w3.index or "C" not in w3.index or not np.isclose(w3.sum(), 1.0):
        raise SystemExit("Rebalancer logic failed")


def _check_portfolio() -> None:
    """Ensure ``Portfolio.rebalance`` records weight history."""
    pf = Portfolio()
    pf.rebalance(
        "2020-01-31",
        pd.DataFrame({"weight": [0.6, 0.4]}, index=["A", "B"]),
    )
    pf.rebalance("2020-02-29", pd.Series({"A": 0.5, "B": 0.5}))
    if len(pf.history) != 2:
        raise SystemExit("Portfolio history length mismatch")
    if not np.isclose(pf.history["2020-02-29"]["A"], 0.5):
        raise SystemExit("Portfolio rebalance values incorrect")


def _check_load_csv_error() -> None:
    """Ensure ``load_csv`` gracefully handles invalid input."""
    if load_csv("_no_such_file_.csv") is not None:
        raise SystemExit("load_csv error handling failed")

    tmp_no_date = Path("demo/exports/no_date.csv")
    tmp_no_date.write_text("A,B\n1,2\n", encoding="utf-8")
    try:
        if load_csv(str(tmp_no_date)) is not None:
            raise SystemExit("load_csv Date-missing check failed")
    finally:
        tmp_no_date.unlink(missing_ok=True)

    tmp_bad = Path("demo/exports/bad.csv")
    tmp_bad.write_text('"a","b"\n1,"2', encoding="utf-8")
    try:
        if load_csv(str(tmp_bad)) is not None:
            raise SystemExit("load_csv parser error check failed")
    finally:
        tmp_bad.unlink(missing_ok=True)


def _check_identify_rf_none() -> None:
    """Ensure ``identify_risk_free_fund`` handles non-numeric data."""
    df = pd.DataFrame({"Date": ["2020-01-01"], "A": ["x"]})
    if identify_risk_free_fund(df) is not None:
        raise SystemExit("identify_risk_free_fund should return None")


def _check_metrics_basic() -> None:
    """Run basic metrics on a short series."""
    s = pd.Series([0.0, 0.01, -0.02])
    bench = pd.Series([0.0, 0.005, -0.01])

    df = pd.DataFrame({"A": s, "B": s * 1.1})
    bench_df = pd.DataFrame({"A": bench, "B": bench * 1.2})

    if not isinstance(metrics.annual_return(s), float):
        raise SystemExit("metrics.annual_return failed")
    if not isinstance(metrics.volatility(s), float):
        raise SystemExit("metrics.volatility failed")
    if not isinstance(metrics.sharpe_ratio(s), float):
        raise SystemExit("metrics.sharpe_ratio failed")
    if not isinstance(metrics.sortino_ratio(s), float):
        raise SystemExit("metrics.sortino_ratio failed")
    if not isinstance(metrics.information_ratio(s, bench), float):
        raise SystemExit("metrics.information_ratio failed")
    if not isinstance(metrics.max_drawdown(s), float):
        raise SystemExit("metrics.max_drawdown failed")
    if not isinstance(metrics.annual_return(df).iloc[0], float):
        raise SystemExit("DataFrame metric annual_return failed")
    if not isinstance(metrics.volatility(df).iloc[0], float):
        raise SystemExit("DataFrame metric volatility failed")
    if not isinstance(metrics.sharpe_ratio(df).iloc[0], float):
        raise SystemExit("DataFrame metric sharpe_ratio failed")
    if not isinstance(metrics.sortino_ratio(df).iloc[0], float):
        raise SystemExit("DataFrame metric sortino_ratio failed")
    if not isinstance(metrics.information_ratio(df, bench_df).iloc[0], float):
        raise SystemExit("DataFrame metric information_ratio failed")
    if not isinstance(metrics.max_drawdown(df).iloc[0], float):
        raise SystemExit("DataFrame metric max_drawdown failed")
    # legacy aliases
    if metrics.annualize_return(s) != metrics.annual_return(s):
        raise SystemExit("annualize_return mismatch")
    if metrics.annualize_volatility(s) != metrics.volatility(s):
        raise SystemExit("annualize_volatility mismatch")
    if metrics.info_ratio(s, bench) != metrics.information_ratio(s, bench):
        raise SystemExit("info_ratio mismatch")


def _check_builtin_metric_aliases() -> None:
    """Ensure legacy metrics are accessible via builtins."""
    import builtins
    import importlib

    legacy = importlib.import_module("tests.legacy_metrics")
    s = pd.Series([0.0, 0.02, -0.01])

    if not hasattr(builtins, "annualize_return"):
        raise SystemExit("builtins missing annualize_return")
    if not hasattr(builtins, "annualize_volatility"):
        raise SystemExit("builtins missing annualize_volatility")

    if builtins.annualize_return(s) != legacy.annualize_return(s):
        raise SystemExit("builtins annualize_return mismatch")
    av = getattr(builtins, "annualize_volatility")
    if av(s) != legacy.annualize_volatility(s):
        raise SystemExit("builtins annualize_volatility mismatch")

    # additional alias checks
    import trend_analysis.metrics as m

    ir1 = m.info_ratio(s, s)
    ir2 = m.information_ratio(s, s)

    def _to_float(x: Any) -> float:
        try:
            return float(x)
        except Exception:  # pragma: no cover - defensive
            return float("nan")

    ir1_f = _to_float(ir1)
    ir2_f = _to_float(ir2)
    if not (np.isnan(ir1_f) and np.isnan(ir2_f)) and ir1_f != ir2_f:
        raise SystemExit("info_ratio alias mismatch")
    if _to_float(m.annualize_sharpe_ratio(s)) != _to_float(m.sharpe_ratio(s)):
        raise SystemExit("annualize_sharpe_ratio alias mismatch")
    if _to_float(m.annualize_sortino_ratio(s)) != _to_float(m.sortino_ratio(s)):
        raise SystemExit("annualize_sortino_ratio alias mismatch")


def _check_metric_helpers() -> None:
    """Exercise internal metric helper functions."""

    ser = pd.Series([0.1, 0.2, -0.1])
    df = pd.DataFrame({"A": ser, "B": ser})

    if not np.isnan(metrics._empty_like(ser, "foo")):
        raise SystemExit("_empty_like Series mismatch")
    empty_df = metrics._empty_like(df, "bar")
    if (
        not isinstance(empty_df, pd.Series)
        or empty_df.name != "bar"
        or not empty_df.isna().all()
    ):
        raise SystemExit("_empty_like DataFrame mismatch")

    try:
        metrics._validate_input([1, 2, 3])  # type: ignore[arg-type]
    except TypeError:
        pass
    else:
        raise SystemExit("_validate_input failed")

    try:
        metrics._check_shapes(ser, df, "demo")
    except ValueError:
        pass
    else:
        raise SystemExit("_check_shapes did not raise")


def _check_selector_errors() -> None:
    """Ensure selectors raise ``KeyError`` for missing columns."""
    df = pd.DataFrame({"A": [1, 2, 3]})

    try:
        RankSelector(top_n=1, rank_column="B").select(df)
    except KeyError:
        pass
    else:
        raise SystemExit("RankSelector missing-column check failed")

    try:
        ZScoreSelector(threshold=0.0).select(df)
    except KeyError:
        pass
    else:
        raise SystemExit("ZScoreSelector missing-column check failed")


def _check_zscore_direction() -> None:
    """Verify ``direction`` parameter handles negative thresholds."""
    df = pd.DataFrame({"metric": [0.0, 1.0, 2.0]}, index=["A", "B", "C"])
    sel = ZScoreSelector(threshold=0.0, direction=-1, column="metric")
    selected, _ = sel.select(df)
    if selected.index.tolist() != ["A"]:
        raise SystemExit("ZScoreSelector direction handling failed")


def _check_weighting_errors() -> None:
    """Ensure weighting classes validate input columns."""
    df = pd.DataFrame({"metric": [0.1, 0.2]}, index=["A", "B"])

    for cls in (ScorePropSimple, ScorePropBayesian):
        try:
            cls("other").weight(df)
        except KeyError:
            continue
        raise SystemExit(f"{cls.__name__} missing-column check failed")


def _check_weighting_zero_sum() -> None:
    """Verify weighting fallbacks when scores sum to zero."""

    df_neg = pd.DataFrame({"Sharpe": [-1.0, -2.0]}, index=["A", "B"])
    out_simple = ScorePropSimple().weight(df_neg)
    if not np.allclose(out_simple["weight"], [0.5, 0.5]):
        raise SystemExit("ScorePropSimple zero-sum fallback failed")
    out_bayes = ScorePropBayesian().weight(df_neg)
    if not np.allclose(out_bayes["weight"], [0.5, 0.5]):
        raise SystemExit("ScorePropBayesian zero-sum fallback failed")


def _check_equal_weight_empty() -> None:
    """Ensure EqualWeight handles empty DataFrames."""

    empty = pd.DataFrame(columns=["Sharpe"])
    out = EqualWeight().weight(empty)
    if not out.empty:
        raise SystemExit("EqualWeight empty DataFrame not handled")


def _check_abw_edge_cases() -> None:
    """Exercise AdaptiveBayesWeighting edge conditions."""

    # mismatched prior_mean length should raise
    try:
        abw_bad = AdaptiveBayesWeighting(prior_mean=np.array([0.5]))
        abw_bad._ensure_index(pd.Index(["A", "B"]))
    except ValueError:
        pass
    else:  # pragma: no cover - should not happen
        raise SystemExit("prior_mean length mismatch not detected")

    # half_life=0 should leave tau equal to obs_tau after updates
    abw = AdaptiveBayesWeighting(max_w=None, half_life=0)
    abw.update(pd.Series({"A": 0.1}), days=30)
    if abw.tau is None or not np.allclose(abw.tau.values, abw.obs_tau):
        raise SystemExit("half_life=0 did not produce obs_tau")
    # expand state when a new fund appears
    abw.update(pd.Series({"A": 0.2, "B": 0.3}), days=30)
    if "B" not in abw.mean.index:
        raise SystemExit("ABW did not add new fund")


def _check_engine_error(cfg: Config) -> None:
    """Ensure ``run_mp`` raises ``FileNotFoundError`` when data is missing."""
    bad = copy.deepcopy(cfg.model_dump())
    bad["data"]["csv_path"] = "_missing.csv"
    try:
        run_mp(Config(**bad))
    except FileNotFoundError:
        pass
    else:  # pragma: no cover - should not happen
        raise SystemExit("run_mp missing-file check failed")


def _check_core_helpers() -> None:
    """Exercise low-level helpers not hit elsewhere."""

    const = pd.Series([0.0, 0.0, 0.0])
    z = rs._zscore(const)
    if not np.allclose(z, 0.0):
        raise SystemExit("_zscore constant input failed")

    try:
        BaseWeighting().weight(pd.DataFrame())
    except NotImplementedError:
        pass
    else:  # pragma: no cover - should not happen
        raise SystemExit("BaseWeighting.weight did not raise")

    alias_cfg = {
        "multi_period": {
            "frequency": "QE",
            "in_sample_len": 4,
            "out_sample_len": 1,
            "start": "2019-01",
            "end": "2020-12",
        }
    }
    periods = scheduler.generate_periods(alias_cfg)
    if not periods or not periods[0].in_start.startswith("2019"):
        raise SystemExit("generate_periods alias handling failed")

    year_cfg = {
        "multi_period": {
            "frequency": "YE",
            "in_sample_len": 1,
            "out_sample_len": 1,
            "start": "2018-01",
            "end": "2024-12",
        }
    }
    year_periods = scheduler.generate_periods(year_cfg)
    if not year_periods or not year_periods[0].in_start.startswith("2018"):
        raise SystemExit("generate_periods YE handling failed")

    df = pd.DataFrame({"A": [0.1, 0.2, 0.3], "B": [0.2, 0.1, 0.4]})
    bs = rs.blended_score(
        df,
        {"Sharpe": 0.6, "AnnualReturn": 0.4},
        RiskStatsConfig(),
    )
    if not isinstance(bs, pd.Series) or len(bs) != 2:
        raise SystemExit("blended_score failed")

    # error handling branches
    try:
        rs.blended_score(df, {}, RiskStatsConfig())
    except ValueError:
        pass
    else:  # pragma: no cover - should not happen
        raise SystemExit("blended_score accepted empty weights")

    try:
        rs._apply_transform(pd.Series([1, 2]), mode="bogus")
    except ValueError:
        pass
    else:  # pragma: no cover - should not happen
        raise SystemExit("_apply_transform invalid mode")

    # verify ddof and window handling
    series = pd.Series([1.0, 2.0, 3.0])
    transformed = rs._apply_transform(series, mode="zscore", window=2, ddof=1)
    if not np.isclose(transformed.iloc[-1], 0.70710678, atol=1e-6):
        raise SystemExit("_apply_transform ddof/window failed")

    try:
        rs._compute_metric_series(df, "NoSuch", RiskStatsConfig())
    except ValueError:
        pass
    else:  # pragma: no cover - should not happen
        raise SystemExit("_compute_metric_series failed to reject unknown metric")


def _check_scheduler_extra() -> None:
    """Exercise additional scheduler edge cases."""

    month_cfg = {
        "multi_period": {
            "frequency": "M",
            "in_sample_len": 2,
            "out_sample_len": 1,
            "start": "2020-01",
            "end": "2020-06",
        }
    }
    months = scheduler.generate_periods(month_cfg)
    if not months or not months[0].in_start.startswith("2020"):
        raise SystemExit("generate_periods M alias failed")

    bad_cfg = {
        "multi_period": {
            "frequency": "BAD",
            "in_sample_len": 1,
            "out_sample_len": 1,
            "start": "2020-01",
            "end": "2020-12",
        }
    }
    try:
        scheduler.generate_periods(bad_cfg)
    except KeyError:
        pass
    else:  # pragma: no cover - should not happen
        raise SystemExit("generate_periods invalid frequency not raised")


def _check_rank_metric_registration() -> None:
    """Ensure ``rank_selection.register_metric`` works end-to-end."""

    @rs.register_metric("DemoAvg")  # type: ignore[misc]
    def _demo_avg(series: pd.Series, **_: object) -> float:
        return float(series.mean())

    df = pd.DataFrame({"A": [0.1, 0.2, 0.3]})
    scores = rs._compute_metric_series(df, "DemoAvg", RiskStatsConfig())
    if not np.isclose(scores["A"], 0.2, atol=1e-9):
        raise SystemExit("register_metric failed")
    rs.METRIC_REGISTRY.pop("DemoAvg", None)


def _check_constants() -> None:
    """Validate key constant values to catch accidental changes."""

    if rs.ASCENDING_METRICS != {"MaxDrawdown"}:
        raise SystemExit("ASCENDING_METRICS unexpected contents")
    if rs.DEFAULT_METRIC != "AnnualReturn":
        raise SystemExit("DEFAULT_METRIC unexpected value")


def _check_abw_halflife() -> None:
    """Verify ``AdaptiveBayesWeighting`` handles ``half_life=0``."""

    abw = AdaptiveBayesWeighting(half_life=0, max_w=None)
    scores = pd.Series({"A": 0.1, "B": 0.2})
    abw.update(scores, 30)
    if abw.tau is None or not np.allclose(abw.tau.values, abw.obs_tau):
        raise SystemExit("ABW half_life=0 update failed")
    abw.update(pd.Series({"A": 0.3, "B": 0.4}), 30)
    if abw.tau is None or not np.allclose(abw.tau.values, abw.obs_tau):
        raise SystemExit("ABW half_life=0 second update failed")
    w = abw.weight(scores.to_frame("z"))
    if not np.isclose(w["weight"].sum(), 1.0):
        raise SystemExit("ABW weight sum mismatch")


def _check_stats_dataclass() -> None:
    """Ensure the ``_Stats`` dataclass exposes all expected fields."""

    expected = {
        "cagr",
        "vol",
        "sharpe",
        "sortino",
        "information_ratio",
        "max_drawdown",
    }
    actual = {f.name for f in fields(pipeline._Stats)}
    if actual != expected:
        raise SystemExit("_Stats fields mismatch")


def _check_notebook_utils() -> None:
    """Exercise notebook helper scripts."""
    src = Path("Vol_Adj_Trend_Analysis1.5.TrEx.ipynb")
    if not src.exists():
        return
    tmp = Path("demo/exports/strip_tmp.ipynb")
    shutil.copy(src, tmp)
    subprocess.run(
        [sys.executable, "tools/strip_output.py", str(tmp)], check=True, shell=False
    )
    data = tmp.read_text(encoding="utf-8")
    if '"outputs": []' not in data:
        raise SystemExit("strip_output failed")
    tmp.unlink()
    subprocess.run(["sh", "tools/pre-commit"], check=True, shell=False)


_check_generate_demo()

_check_generate_demo_help()


cfg = load("config/demo.yml")
regime_cfg = getattr(cfg, "regime", {})
demo_df = _check_demo_data(cfg)
if cfg.export.get("filename") != "alias_demo.csv":
    raise SystemExit("Output alias not parsed")
results = run_mp(cfg)
num_periods = len(results)
periods = scheduler.generate_periods(cfg.model_dump())
expected_periods = len(periods)
print(f"Generated {num_periods} period results (expected {expected_periods})")
if num_periods != expected_periods:
    raise SystemExit("Multi-period demo produced an unexpected number of periods")
if num_periods <= 1:
    raise SystemExit("Multi-period demo produced insufficient results")

df_pre = demo_df
results_pre = run_mp(cfg, df_pre)
if len(results_pre) != num_periods:
    raise SystemExit("Preloaded DataFrame run mismatch")

# check that the generated periods line up with the scheduler output
result_periods = [r["period"] for r in results]
sched_tuples = [(p.in_start, p.in_end, p.out_start, p.out_end) for p in periods]
if result_periods != sched_tuples:
    raise SystemExit("Period sequence mismatch")

score_frames = {r["period"][3]: r["score_frame"] for r in results}

# export all score frames in one go so Excel, CSV, JSON and TXT versions are
# produced for CI. Each period becomes a separate sheet/file.
mp_prefix = Path("demo/exports/multi_period_scores")
export.export_data(score_frames, str(mp_prefix), formats=["xlsx", "csv", "json", "txt"])
if not mp_prefix.with_suffix(".xlsx").exists():
    raise SystemExit("Multi-period score frame export failed")

# ensure metadata lines up with the generated periods
for r in results:
    sf = r["score_frame"]
    p = r["period"]
    if sf.attrs.get("period") != (p[0][:7], p[1][:7]):
        raise SystemExit("Score frame period metadata mismatch")
    if sf.attrs.get("insample_len", 0) <= 0:
        raise SystemExit("Score frame contains no data")

# Exercise multi-period export helpers
frames = export.workbook_frames_from_results(results)
if "summary" not in frames:
    raise SystemExit("workbook_frames_from_results missing summary")
results_prefix = Path("demo/exports/workbook_frames")
export.export_data(frames, str(results_prefix), formats=["xlsx", "csv", "json", "txt"])
if not results_prefix.with_suffix(".xlsx").exists():
    raise SystemExit("Workbook frame export failed")
csv_files = list(results_prefix.parent.glob(f"{results_prefix.stem}_*.csv"))
json_files = list(results_prefix.parent.glob(f"{results_prefix.stem}_*.json"))
txt_files = list(results_prefix.parent.glob(f"{results_prefix.stem}_*.txt"))
if not csv_files or not json_files or not txt_files:
    raise SystemExit("Workbook frame CSV/JSON/TXT missing")
phase1_prefix = Path("demo/exports/phase1_multi")
export.export_phase1_multi_metrics(
    results,
    str(phase1_prefix),
    formats=["xlsx", "csv", "json", "txt"],
    include_metrics=True,
)
if not phase1_prefix.with_suffix(".xlsx").exists():
    raise SystemExit("Phase1 multi metrics export failed")
if not phase1_prefix.with_name(f"{phase1_prefix.stem}_periods.csv").exists():
    raise SystemExit("Phase1 multi metrics CSV missing")
if not phase1_prefix.with_name(f"{phase1_prefix.stem}_periods.json").exists():
    raise SystemExit("Phase1 multi metrics JSON missing")
if not phase1_prefix.with_name(f"{phase1_prefix.stem}_summary.csv").exists():
    raise SystemExit("Phase1 multi metrics summary CSV missing")
if not phase1_prefix.with_name(f"{phase1_prefix.stem}_summary.json").exists():
    raise SystemExit("Phase1 multi metrics summary JSON missing")
if not phase1_prefix.with_name(f"{phase1_prefix.stem}_periods.txt").exists():
    raise SystemExit("Phase1 multi metrics TXT missing")
if not phase1_prefix.with_name(f"{phase1_prefix.stem}_summary.txt").exists():
    raise SystemExit("Phase1 multi metrics summary TXT missing")
if not phase1_prefix.with_name(f"{phase1_prefix.stem}_metrics.csv").exists():
    raise SystemExit("Phase1 multi metrics metrics CSV missing")
if not phase1_prefix.with_name(f"{phase1_prefix.stem}_metrics.json").exists():
    raise SystemExit("Phase1 multi metrics metrics JSON missing")
if not phase1_prefix.with_name(f"{phase1_prefix.stem}_metrics_summary.csv").exists():
    raise SystemExit("Phase1 multi metrics metrics summary CSV missing")
if not phase1_prefix.with_name(f"{phase1_prefix.stem}_metrics_summary.json").exists():
    raise SystemExit("Phase1 multi metrics metrics summary JSON missing")
    if not phase1_prefix.with_name(
        f"{phase1_prefix.stem}_metrics_summary.txt"
    ).exists():
        raise SystemExit("Phase1 multi metrics metrics summary TXT missing")

# Additional OS summaries and churn report for convenience
_out_dir = Path("demo/exports")
_out_dir.mkdir(parents=True, exist_ok=True)

# Per-period EW OS stats
_period_rows = []
for _res in results:
    _p = _res.get("period")
    _label = str(_p[3]) if isinstance(_p, (list, tuple)) and len(_p) >= 4 else str(_p)
    _stats = _res.get("out_ew_stats")
    if _stats is not None:
        _period_rows.append(
            {
                "period": _label,
                "cagr": getattr(_stats, "cagr", float("nan")),
                "vol": getattr(_stats, "vol", float("nan")),
                "sharpe": getattr(_stats, "sharpe", float("nan")),
                "sortino": getattr(_stats, "sortino", float("nan")),
                "information_ratio": getattr(_stats, "information_ratio", float("nan")),
                "max_drawdown": getattr(_stats, "max_drawdown", float("nan")),
            }
        )
pd.DataFrame(_period_rows).to_csv(_out_dir / "period_os_stats.csv", index=False)

# Combined EW OS stats (single row)
_combined = export.combined_summary_result(results)
_comb_stats = _combined.get("out_ew_stats")
if _comb_stats is not None:
    pd.DataFrame(
        [
            {  # type: ignore[list-item]
                "period": "all_periods",
                "cagr": getattr(_comb_stats, "cagr", float("nan")),
                "vol": getattr(_comb_stats, "vol", float("nan")),
                "sharpe": getattr(_comb_stats, "sharpe", float("nan")),
                "sortino": getattr(_comb_stats, "sortino", float("nan")),
                "information_ratio": getattr(
                    _comb_stats, "information_ratio", float("nan")
                ),
                "max_drawdown": getattr(_comb_stats, "max_drawdown", float("nan")),
            }
        ]
    ).to_csv(_out_dir / "combined_os_stats.csv", index=False)

# Portfolio churn (entries/exits per period)
_prev: set[str] | None = None
_churn_rows = []
for _res in results:
    _p = _res.get("period")
    _label = str(_p[3]) if isinstance(_p, (list, tuple)) and len(_p) >= 4 else str(_p)
    _sel = set(_res.get("selected_funds", []))
    _entries: list[str] = []
    _exits: list[str] = []
    if _prev is not None:
        _entries = sorted(_sel - _prev)
        _exits = sorted(_prev - _sel)
    _churn_rows.append(
        {
            "period": _label,
            "selected_funds": ",".join(sorted(_sel)),
            "entries": ",".join(_entries),
            "exits": ",".join(_exits),
        }
    )
    _prev = _sel
pd.DataFrame(_churn_rows).to_csv(_out_dir / "portfolio_churn.csv", index=False)
mpm_prefix = Path("demo/exports/multi_period_metrics")
export.export_multi_period_metrics(
    results,
    str(mpm_prefix),
    formats=["xlsx", "csv", "json", "txt"],
    include_metrics=True,
)
if not mpm_prefix.with_suffix(".xlsx").exists():
    raise SystemExit("Multi-period metrics export failed")
if not mpm_prefix.with_name(f"{mpm_prefix.stem}_periods.csv").exists():
    raise SystemExit("Multi-period metrics CSV missing")
if not mpm_prefix.with_name(f"{mpm_prefix.stem}_periods.json").exists():
    raise SystemExit("Multi-period metrics JSON missing")
if not mpm_prefix.with_name(f"{mpm_prefix.stem}_periods.txt").exists():
    raise SystemExit("Multi-period metrics TXT missing")
if not mpm_prefix.with_name(f"{mpm_prefix.stem}_summary.csv").exists():
    raise SystemExit("Multi-period metrics summary CSV missing")
if not mpm_prefix.with_name(f"{mpm_prefix.stem}_summary.json").exists():
    raise SystemExit("Multi-period metrics summary JSON missing")
if not mpm_prefix.with_name(f"{mpm_prefix.stem}_summary.txt").exists():
    raise SystemExit("Multi-period metrics summary TXT missing")
if not mpm_prefix.with_name(f"{mpm_prefix.stem}_metrics.csv").exists():
    raise SystemExit("Multi-period metrics metrics CSV missing")
if not mpm_prefix.with_name(f"{mpm_prefix.stem}_metrics.json").exists():
    raise SystemExit("Multi-period metrics metrics JSON missing")
if not mpm_prefix.with_name(f"{mpm_prefix.stem}_metrics.txt").exists():
    raise SystemExit("Multi-period metrics metrics TXT missing")
if not mpm_prefix.with_name(f"{mpm_prefix.stem}_metrics_summary.csv").exists():
    raise SystemExit("Multi-period metrics metrics summary CSV missing")
if not mpm_prefix.with_name(f"{mpm_prefix.stem}_metrics_summary.json").exists():
    raise SystemExit("Multi-period metrics metrics summary JSON missing")
if not mpm_prefix.with_name(f"{mpm_prefix.stem}_metrics_summary.txt").exists():
    raise SystemExit("Multi-period metrics metrics summary TXT missing")

wb_mpm = openpyxl.load_workbook(mpm_prefix.with_suffix(".xlsx"))
expected = {str(r["period"][3]) for r in results}
expected.add("summary")
metric_tabs = {f"metrics_{s}" for s in expected if s != "summary"}
metric_tabs.add("metrics_summary")
all_tabs = expected.union(metric_tabs)
if not all_tabs.issubset(set(wb_mpm.sheetnames)):
    raise SystemExit("Multi-period metrics workbook sheets mismatch")

# Verify exporters also handle the no-metrics case
phase1_nom_prefix = Path("demo/exports/phase1_multi_nom")
export.export_phase1_multi_metrics(
    results,
    str(phase1_nom_prefix),
    formats=["xlsx", "csv", "json", "txt"],
    include_metrics=False,
)
if not phase1_nom_prefix.with_suffix(".xlsx").exists():
    raise SystemExit("Phase1 multi metrics nometrics Excel missing")
for ext in ("csv", "json", "txt"):
    if not phase1_nom_prefix.with_name(
        f"{phase1_nom_prefix.stem}_periods.{ext}"
    ).exists():
        raise SystemExit(f"Phase1 multi metrics nometrics {ext} missing")
    if not phase1_nom_prefix.with_name(
        f"{phase1_nom_prefix.stem}_summary.{ext}"
    ).exists():
        raise SystemExit(f"Phase1 multi metrics nometrics summary {ext} missing")

mpm_nom_prefix = Path("demo/exports/multi_period_nometrics")
export.export_multi_period_metrics(
    results,
    str(mpm_nom_prefix),
    formats=["xlsx", "csv", "json", "txt"],
    include_metrics=False,
)
if not mpm_nom_prefix.with_suffix(".xlsx").exists():
    raise SystemExit("Multi-period metrics nometrics export failed")
for ext in ("csv", "json", "txt"):
    if not mpm_nom_prefix.with_name(f"{mpm_nom_prefix.stem}_periods.{ext}").exists():
        raise SystemExit(f"Multi-period metrics nometrics {ext} missing")
    if not mpm_nom_prefix.with_name(f"{mpm_nom_prefix.stem}_summary.{ext}").exists():
        raise SystemExit(f"Multi-period metrics nometrics summary {ext} missing")
wb_nom = openpyxl.load_workbook(mpm_nom_prefix.with_suffix(".xlsx"))
if any(name.startswith("metrics_") for name in wb_nom.sheetnames):
    raise SystemExit("Nometrics workbook contains metrics sheets")
wb_direct = Path("demo/exports/phase1_direct.xlsx")
export.export_phase1_workbook(results, str(wb_direct))
if not wb_direct.exists():
    raise SystemExit("export_phase1_workbook failed")
wb = openpyxl.load_workbook(wb_direct)
# Phase-1 workbook now (by design) includes an "execution_metrics" sheet
# alongside one sheet per period and the leading "summary" sheet. Keep the
# demo expectation flexible: all required sheets must be present, but the
# exporter may add new diagnostic sheets over time without breaking CI.
period_sheet_names = {str(r["period"][3]) for r in results}
required = {"summary", "execution_metrics"} | period_sheet_names
missing = required.difference(wb.sheetnames)
if missing:
    raise SystemExit(f"phase1 workbook sheets mismatch (missing: {sorted(missing)})")
pf_frames = export.period_frames_from_results(results)
if len(pf_frames) != len(results):
    raise SystemExit("period_frames_from_results count mismatch")
if "summary" in pf_frames:
    raise SystemExit("period_frames_from_results should not include summary")
pf_prefix = Path("demo/exports/period_frames")
export.export_data(pf_frames, str(pf_prefix), formats=["xlsx", "csv", "json", "txt"])
if not pf_prefix.with_suffix(".xlsx").exists():
    raise SystemExit("period frame Excel missing")
created = list(pf_prefix.parent.glob(f"{pf_prefix.stem}_*.csv"))
if not created:
    raise SystemExit("period frame CSV/JSON/TXT missing")
summary = export.combined_summary_result(results)
summary_frame = export.summary_frame_from_result(summary)
metrics_frame = export.metrics_from_result(summary)
if summary_frame.empty or metrics_frame.empty:
    raise SystemExit("Summary export helpers failed")
summ_prefix = Path("demo/exports/summary_frames")
export.export_data(
    {"summary": summary_frame, "metrics": metrics_frame},
    str(summ_prefix),
    formats=["xlsx", "csv", "json", "txt"],
)
if not summ_prefix.with_suffix(".xlsx").exists():
    raise SystemExit("summary frame Excel missing")

# Export per-period metrics using all exporters to cover the helper
metrics_prefix = Path("demo/exports/period_metrics")
period_metrics = {
    str(res["period"][3]): export.metrics_from_result(res) for res in results
}
export.export_data(
    period_metrics, str(metrics_prefix), formats=["xlsx", "csv", "json", "txt"]
)
if not metrics_prefix.with_suffix(".xlsx").exists():
    raise SystemExit("period metrics Excel missing")
created = list(metrics_prefix.parent.glob(f"{metrics_prefix.stem}_*.csv"))
if not created:
    raise SystemExit("period metrics CSV/JSON/TXT missing")

# Verify case-insensitive exporter dispatch
mixed_prefix = Path("demo/exports/mixed_case")
export.export_data(frames, str(mixed_prefix), formats=["CSV", "Excel", "Json", "TXT"])
if not mixed_prefix.with_suffix(".xlsx").exists():
    raise SystemExit("Mixed-case Excel export failed")
for ext in ("csv", "json", "txt"):
    if not list(mixed_prefix.parent.glob(f"{mixed_prefix.stem}_*.{ext}")):
        raise SystemExit(f"Mixed-case {ext} export missing")

# Exercise rank_select_funds via the additional inclusion approaches
df_full = demo_df
rf_col = identify_risk_free_fund(df_full)
if rf_col is None:
    raise SystemExit("identify_risk_free_fund failed")
mask = df_full["Date"].between(cfg.sample_split["in_start"], cfg.sample_split["in_end"])
window = df_full.loc[mask].drop(columns=["Date"])
rs_cfg = RiskStatsConfig()
top_pct_ids = rank_select_funds(
    window,
    rs_cfg,
    inclusion_approach="top_pct",
    pct=0.2,
    score_by="Sharpe",
)
if not top_pct_ids:
    raise SystemExit("top_pct selection produced no funds")
threshold_ids = rank_select_funds(
    window,
    rs_cfg,
    inclusion_approach="threshold",
    threshold=0.0,
    score_by="Sharpe",
)
if not threshold_ids:
    raise SystemExit("threshold selection produced no funds")
blended_ids = rank_select_funds(
    window,
    rs_cfg,
    inclusion_approach="top_n",
    n=3,
    score_by="blended",
    blended_weights={"Sharpe": 0.6, "AnnualReturn": 0.3, "MaxDrawdown": 0.1},
)
if not blended_ids:
    raise SystemExit("blended selection produced no funds")
zscore_ids = rank_select_funds(
    window,
    rs_cfg,
    inclusion_approach="top_n",
    n=2,
    score_by="Sharpe",
    transform="zscore",
)
if not zscore_ids:
    raise SystemExit("zscore selection produced no funds")

zwin_ids = rank_select_funds(
    window,
    rs_cfg,
    inclusion_approach="top_n",
    n=2,
    score_by="Sharpe",
    transform="zscore",
    zscore_window=5,
)
if not zwin_ids:
    raise SystemExit("zscore window selection produced no funds")

percentile_ids = rank_select_funds(
    window,
    rs_cfg,
    inclusion_approach="top_pct",
    pct=0.3,
    score_by="Sharpe",
    transform="percentile",
    rank_pct=0.5,
)
if not percentile_ids:
    raise SystemExit("percentile transform produced no funds")

rank_ids = rank_select_funds(
    window,
    rs_cfg,
    inclusion_approach="top_n",
    n=3,
    score_by="Sharpe",
    transform="rank",
)
if not rank_ids:
    raise SystemExit("rank transform produced no funds")

alias_ids = rank_select_funds(
    window,
    rs_cfg,
    inclusion_approach="top_n",
    n=2,
    score_by="Sharpe",
    transform_mode="zscore",
)
if not alias_ids:
    raise SystemExit("transform_mode alias failed")

# verify ranking works with an ascending metric alias
maxdd_ids = rank_select_funds(
    window,
    rs_cfg,
    inclusion_approach="top_n",
    n=2,
    score_by="MaxDrawdown",
)
if not maxdd_ids:
    raise SystemExit("max_drawdown selection produced no funds")

# verify ranking with the InformationRatio metric works
ir_ids = rank_select_funds(
    window,
    rs_cfg,
    inclusion_approach="top_n",
    n=2,
    score_by="InformationRatio",
)
if not ir_ids:
    raise SystemExit("information_ratio selection produced no funds")

df_tmp = pd.DataFrame({"Time": ["2020-01-01", "2020-02-01"], "Val": [0.0, 0.1]})
dt_tmp = ensure_datetime(df_tmp, "Time")
if not pd.api.types.is_datetime64_any_dtype(dt_tmp["Time"]):
    raise SystemExit("ensure_datetime custom column failed")
ew_df = EqualWeight().weight(
    pd.DataFrame({"metric": [1.0, 2.0, 3.0]}, index=["A", "B", "C"])
)
if not np.isclose(ew_df["weight"].sum(), 1.0, atol=1e-4):
    raise SystemExit("EqualWeight weight sum mismatch")

direct_res = pipeline._run_analysis(
    df_full,
    str(cfg.sample_split["in_start"]),
    str(cfg.sample_split["in_end"]),
    str(cfg.sample_split["out_start"]),
    str(cfg.sample_split["out_end"]),
    cfg.vol_adjust.get("target_vol", 1.0),
    getattr(cfg, "run", {}).get("monthly_cost", 0.0),
    selection_mode="rank",
    rank_kwargs={"inclusion_approach": "top_n", "n": 2, "score_by": "Sharpe"},
)
if direct_res is None or "score_frame" not in direct_res:
    raise SystemExit("_run_analysis direct call failed")

# cover helper with missing type annotations
scores = rs._compute_metric_series(window, "Sharpe", rs_cfg)
extra_ids = rs.some_function_missing_annotation(
    scores,
    "top_n",
    n=2,
    ascending=False,
)
if len(extra_ids) != 2:
    raise SystemExit("some_function_missing_annotation failed")
pct_ids = rs.some_function_missing_annotation(
    scores,
    "top_pct",
    pct=0.5,
    ascending=False,
)
if not pct_ids:
    raise SystemExit("some_function_missing_annotation top_pct failed")
thr_ids = rs.some_function_missing_annotation(
    scores,
    "threshold",
    threshold=0.0,
    ascending=False,
)
if not thr_ids:
    raise SystemExit("some_function_missing_annotation threshold failed")

# also cover the ascending=True branch for completeness
asc_ids = rs.some_function_missing_annotation(
    scores,
    "threshold",
    threshold=scores.max(),
    ascending=True,
)
if not asc_ids:
    raise SystemExit("some_function_missing_annotation ascending branch failed")

# quality_filter and select_funds interfaces
qcfg = rs.FundSelectionConfig(max_missing_ratio=0.5)
eligible = rs.quality_filter(df_full, qcfg)
if not eligible or not set(eligible).issubset(df_full.columns):
    raise SystemExit("quality_filter failed")

filtered = rs._quality_filter(
    df_full,
    [c for c in df_full.columns if c not in {"Date", rf_col}],
    str(cfg.sample_split["in_start"]),
    str(cfg.sample_split["out_end"]),
    qcfg,
)
if not filtered:
    raise SystemExit("_quality_filter returned no funds")

simple_sel = rs.select_funds(df_full, rf_col, mode="random", n=2)
if len(simple_sel) != 2:
    raise SystemExit("select_funds simple mode failed")

cols = [c for c in df_full.columns if c not in {"Date", rf_col}]
ext_sel = rs.select_funds(
    df_full,
    rf_col,
    cols,
    str(cfg.sample_split["in_start"]),
    str(cfg.sample_split["in_end"]),
    str(cfg.sample_split["out_start"]),
    str(cfg.sample_split["out_end"]),
    qcfg,
    "rank",
    2,
    rank_kwargs={"inclusion_approach": "top_n", "n": 2, "score_by": "Sharpe"},
)
if len(ext_sel) != 2:
    raise SystemExit("select_funds extended mode failed")

ext_sel_direct = rs.select_funds_extended(
    df_full,
    rf_col,
    cols,
    str(cfg.sample_split["in_start"]),
    str(cfg.sample_split["in_end"]),
    str(cfg.sample_split["out_start"]),
    str(cfg.sample_split["out_end"]),
    qcfg,
    selection_mode="rank",
    rank_kwargs={"inclusion_approach": "top_n", "n": 2, "score_by": "Sharpe"},
)
if len(ext_sel_direct) != 2:
    raise SystemExit("select_funds_extended direct call failed")

abw = AdaptiveBayesWeighting(max_w=None)
pf_abw = _check_schedule(
    score_frames,
    RankSelector(top_n=3, rank_column="Sharpe"),
    abw,
    cfg,
    rank_column="Sharpe",
)
abw_prefix = Path("demo/exports/abw_weights")
export.export_data(
    {"weights": pd.DataFrame(pf_abw.history).T},
    str(abw_prefix),
    formats=["xlsx", "csv", "json", "txt"],
)
if not abw_prefix.with_suffix(".xlsx").exists():
    raise SystemExit("ABW weight export failed")
for ext in ("csv", "json", "txt"):
    if not abw_prefix.with_name(f"{abw_prefix.stem}_weights.{ext}").exists():
        raise SystemExit(f"ABW weight {ext} missing")
state = abw.get_state()
abw2 = AdaptiveBayesWeighting(max_w=None)
abw2.set_state(state)
if abw2.get_state() != state:
    raise SystemExit("AdaptiveBayesWeighting state mismatch")

_check_schedule(
    score_frames,
    ZScoreSelector(threshold=0.0, column="Sharpe"),
    EqualWeight(),
    cfg,
    rank_column="Sharpe",
)

_check_schedule(
    score_frames,
    RankSelector(top_n=3, rank_column="Sharpe"),
    ScorePropSimple("Sharpe"),
    cfg,
    rank_column="Sharpe",
)

_check_schedule(
    score_frames,
    RankSelector(top_n=2, rank_column="MaxDrawdown"),
    ScorePropBayesian("Sharpe"),
    cfg,
    rank_column="MaxDrawdown",
)

# Exercise the single-period pipeline and export helpers - using run_full
# to avoid redundant computation
full_res = pipeline.run_full(cfg)
if not isinstance(full_res, dict):
    raise SystemExit("pipeline.run_full did not return a dict")


# Extract metrics DataFrame from full results using helper to avoid code duplication
def extract_metrics_df(full_res):
    stats = full_res.get("out_sample_stats", {})
    if not stats:
        raise SystemExit("pipeline.run_full out_sample_stats missing")
    metrics_df = pd.DataFrame({k: vars(v) for k, v in stats.items()}).T
    for label, ir_map in full_res.get("benchmark_ir", {}).items():
        col = f"ir_{label}"
        metrics_df[col] = pd.Series(
            {
                k: v
                for k, v in ir_map.items()
                if k not in {"equal_weight", "user_weight"}
            }
        )
    return metrics_df


metrics_df = extract_metrics_df(full_res)

if metrics_df.empty:
    raise SystemExit("pipeline.run_full produced empty metrics")
if "ir_spx" not in metrics_df.columns:
    raise SystemExit("pipeline.run_full missing ir_spx column")
expected_cols = {
    "cagr",
    "vol",
    "sharpe",
    "sortino",
    "information_ratio",
    "max_drawdown",
    "ir_spx",
}
if set(metrics_df.columns) != expected_cols:
    raise SystemExit("pipeline.run_full column mismatch")
out_prefix = Path("demo/exports/pipeline_demo")
export.export_data(
    {"metrics": metrics_df},
    str(out_prefix),
    formats=["xlsx", "csv", "json", "txt"],
)
if not out_prefix.with_suffix(".xlsx").exists():
    raise SystemExit("Excel export failed")
if not out_prefix.with_name(f"{out_prefix.stem}_metrics.csv").exists():
    raise SystemExit("CSV export failed")
if not out_prefix.with_name(f"{out_prefix.stem}_metrics.json").exists():
    raise SystemExit("JSON export failed")
if not out_prefix.with_name(f"{out_prefix.stem}_metrics.txt").exists():
    raise SystemExit("TXT export failed")

# Additional validation of full_res content (stats already validated above)
sf = full_res.get("score_frame")
if sf is None or sf.empty:
    raise SystemExit("pipeline.run_full missing score_frame")
b_ir = full_res.get("benchmark_ir", {})
if "spx" not in b_ir or "equal_weight" not in b_ir.get("spx", {}):
    raise SystemExit("pipeline.run_full benchmark_ir missing")
regime_table_full = full_res.get("performance_by_regime")
if not isinstance(regime_table_full, pd.DataFrame) or regime_table_full.empty:
    raise SystemExit("pipeline.run_full missing regime table")
regime_summary_full = str(full_res.get("regime_summary", ""))
regime_notes_full = full_res.get("regime_notes") or []
if not regime_summary_full.strip() and not any(
    str(note).strip() for note in regime_notes_full
):
    raise SystemExit("pipeline.run_full missing regime insight")
_oss = full_res.get("out_sample_stats", {})
_oss = _oss if isinstance(_oss, dict) else {}
for obj in _oss.values():
    if not hasattr(obj, "information_ratio"):
        raise SystemExit("_Stats missing information_ratio")

# Reuse the sample split for the convenience wrapper
split = cfg.sample_split
# Exercise the convenience wrapper around ``_run_analysis``
analysis_res = pipeline.run_analysis(
    df_full,
    str(split.get("in_start")),
    str(split.get("in_end")),
    str(split.get("out_start")),
    str(split.get("out_end")),
    cfg.vol_adjust.get("target_vol", 1.0),
    getattr(cfg, "run", {}).get("monthly_cost", 0.0),
    selection_mode="rank",
    rank_kwargs={"n": 5, "score_by": "Sharpe", "inclusion_approach": "top_n"},
    regime_cfg=regime_cfg,
)
if analysis_res is None or analysis_res.get("score_frame") is None:
    raise SystemExit("pipeline.run_analysis failed")
analysis_table = analysis_res.get("performance_by_regime")
if not isinstance(analysis_table, pd.DataFrame) or analysis_table.empty:
    raise SystemExit("pipeline.run_analysis missing regime table")
analysis_summary = str(analysis_res.get("regime_summary", ""))
analysis_notes = analysis_res.get("regime_notes") or []
if not analysis_summary.strip() and not any(str(n).strip() for n in analysis_notes):
    raise SystemExit("pipeline.run_analysis missing regime insight")
analysis_idx = pipeline.run_analysis(
    df_full,
    str(split.get("in_start")),
    str(split.get("in_end")),
    str(split.get("out_start")),
    str(split.get("out_end")),
    cfg.vol_adjust.get("target_vol", 1.0),
    getattr(cfg, "run", {}).get("monthly_cost", 0.0),
    indices_list=["Mgr_01", "Mgr_02"],
    regime_cfg=regime_cfg,
)
if analysis_idx is None or not analysis_idx.get("benchmark_stats"):
    raise SystemExit("pipeline.run_analysis with indices_list failed")

# Verify custom_weights behaviour using a direct _run_analysis call
cw_res = pipeline._run_analysis(
    df_full[["Date", "Mgr_01", "Mgr_02"]],
    str(split.get("in_start")),
    str(split.get("in_end")),
    str(split.get("out_start")),
    str(split.get("out_end")),
    cfg.vol_adjust.get("target_vol", 1.0),
    getattr(cfg, "run", {}).get("monthly_cost", 0.0),
    selection_mode="all",
    custom_weights={"Mgr_01": 60, "Mgr_02": 40},
)
fw = cw_res.get("fund_weights", {})
expected = {"Mgr_01": 0.6, "Mgr_02": 0.4}
for key, val in expected.items():
    if key in fw and not np.isclose(fw[key], val, atol=1e-6):
        raise SystemExit("custom_weights not applied")

# Export a formatted summary workbook and text summary
split = cfg.sample_split
sheet_fmt = export.make_summary_formatter(
    full_res,
    str(split.get("in_start")),
    str(split.get("in_end")),
    str(split.get("out_start")),
    str(split.get("out_end")),
)
summary_prefix = Path("demo/exports/summary_demo")
export.export_to_excel(
    {"metrics": metrics_df, "summary": pd.DataFrame()},
    str(summary_prefix.with_suffix(".xlsx")),
    default_sheet_formatter=sheet_fmt,
)
text = export.format_summary_text(
    full_res,
    str(split.get("in_start")),
    str(split.get("in_end")),
    str(split.get("out_start")),
    str(split.get("out_end")),
)
if "Vol-Adj Trend Analysis" not in text:
    raise SystemExit("Text summary missing header")
if not summary_prefix.with_suffix(".xlsx").exists():
    raise SystemExit("Summary Excel not created")

# Exercise formatter registry helpers
export.reset_formatters_excel()
if export.FORMATTERS_EXCEL:
    raise SystemExit("Formatter registry not cleared")


@export.register_formatter_excel("dummy")  # type: ignore[misc]
def _demo_fmt(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    wb: openpyxl.Workbook,
) -> None:
    ws.write(0, 0, "demo")


if "dummy" not in export.FORMATTERS_EXCEL:
    raise SystemExit("Formatter registration failed")


dummy_prefix = Path("demo/exports/dummy")
export.export_to_excel(
    {"dummy": pd.DataFrame({"A": [1, 2]})}, str(dummy_prefix.with_suffix(".xlsx"))
)
if not dummy_prefix.with_suffix(".xlsx").exists():
    raise SystemExit("Custom Excel export failed")

# Exercise export_data with a formatter to cover formatting hooks
fmt_prefix = Path("demo/exports/formatted_demo")
df_fmt = pd.DataFrame({"A": [1, 2], "B": [3, 4]})


def _double(df: pd.DataFrame) -> pd.DataFrame:
    return df * 2


export.export_data(
    {"tbl": df_fmt},
    str(fmt_prefix),
    formats=["xlsx", "csv", "json", "txt"],
    formatter=_double,
)
csv_file = fmt_prefix.with_name(f"{fmt_prefix.stem}_tbl.csv")
if not csv_file.exists():
    raise SystemExit("Formatted CSV not created")
chk = pd.read_csv(csv_file)
if chk["A"].iloc[0] != 2:
    raise SystemExit("Formatter did not apply")
txt_file = fmt_prefix.with_name(f"{fmt_prefix.stem}_tbl.txt")
if not txt_file.exists():
    raise SystemExit("Formatted TXT not created")
txt_lines = txt_file.read_text().splitlines()
if len(txt_lines) < 2 or not txt_lines[1].strip().startswith("2"):
    raise SystemExit("TXT formatter did not apply")

# Exercise individual exporters and the period formatter helper
indiv_prefix = Path("demo/exports/indiv")
export.export_to_csv({"metrics": metrics_df}, str(indiv_prefix))
export.export_to_json({"metrics": metrics_df}, str(indiv_prefix))
export.export_to_txt({"metrics": metrics_df}, str(indiv_prefix))
for ext in ("csv", "json", "txt"):
    if not indiv_prefix.with_name(f"{indiv_prefix.stem}_metrics.{ext}").exists():
        raise SystemExit(f"export_to_{ext} failed")

extra_prefix = Path("demo/exports/extra_period.xlsx")
export.reset_formatters_excel()
if export.FORMATTERS_EXCEL:
    raise SystemExit("Formatter registry not cleared before period formatter")
export.make_period_formatter(
    "extra",
    results[0],
    str(results[0]["period"][0]),
    str(results[0]["period"][1]),
    str(results[0]["period"][2]),
    str(results[0]["period"][3]),
)
if "extra" not in export.FORMATTERS_EXCEL:
    raise SystemExit("make_period_formatter did not register formatter")
export.export_to_excel(
    {"extra": export.summary_frame_from_result(results[0])},
    str(extra_prefix),
)
wb_extra = openpyxl.load_workbook(extra_prefix)
if wb_extra["extra"]["A1"].value != "Vol-Adj Trend Analysis":
    raise SystemExit("make_period_formatter formatting failed")
export.reset_formatters_excel()
if export.FORMATTERS_EXCEL:
    raise SystemExit("Formatter registry not cleared after export")

_check_gui("config/demo.yml")
_check_datagrid_override()
_check_plugin_discovery()
_check_selection_modes(cfg)
_check_cli_env("config/demo.yml")
_check_cli_env_multi("config/demo.yml")
_check_cli("config/demo.yml", cfg.data["csv_path"])
_check_misc("config/demo.yml", cfg, results)
_check_config_dump(cfg)
_check_default_load()
_check_rebalancer_logic()
_check_portfolio()
_check_load_csv_error()
_check_identify_rf_none()
_check_metrics_basic()
_check_builtin_metric_aliases()
_check_metric_helpers()
_check_selector_errors()
_check_zscore_direction()
_check_weighting_errors()
_check_weighting_zero_sum()
_check_equal_weight_empty()
_check_abw_edge_cases()
_check_core_helpers()
_check_scheduler_extra()
_check_rank_metric_registration()
_check_abw_halflife()
_check_stats_dataclass()
_check_constants()
_check_engine_error(cfg)
_check_notebook_utils()


def _check_run_analysis_errors(cfg: Config) -> None:
    """Exercise error-handling branches in ``_run_analysis``."""

    res = pipeline._run_analysis(
        None,
        str(cfg.sample_split["in_start"]),
        str(cfg.sample_split["in_end"]),
        str(cfg.sample_split["out_start"]),
        str(cfg.sample_split["out_end"]),
        cfg.vol_adjust.get("target_vol", 1.0),
        getattr(cfg, "run", {}).get("monthly_cost", 0.0),
    )
    if res is not None:
        raise SystemExit("_run_analysis did not return None on missing df")

    try:
        pipeline._run_analysis(
            pd.DataFrame({"A": [0.1, 0.2]}),
            str(cfg.sample_split["in_start"]),
            str(cfg.sample_split["in_end"]),
            str(cfg.sample_split["out_start"]),
            str(cfg.sample_split["out_end"]),
            cfg.vol_adjust.get("target_vol", 1.0),
            getattr(cfg, "run", {}).get("monthly_cost", 0.0),
        )
    except ValueError:
        pass
    else:
        raise SystemExit("_run_analysis accepted DataFrame without Date column")


# ------------------------------------------------------------
# Additional error handling checks


def _check_export_errors() -> None:
    """Ensure exporter functions validate input."""
    df = pd.DataFrame({"A": [1]})
    try:
        export.export_data({"A": df}, "demo/exports/bad", formats=["foo"])
    except ValueError:
        pass
    else:  # pragma: no cover - should not happen
        raise SystemExit("export_data did not reject unknown format")


def _check_config_errors() -> None:
    """Verify config loader rejects invalid YAML."""
    tmp = Path("demo/exports/invalid.yml")
    tmp.write_text("- a\n- b\n", encoding="utf-8")
    try:
        load(str(tmp))
    except TypeError:
        pass
    else:  # pragma: no cover - should not happen
        raise SystemExit("config.load accepted non-mapping YAML")
    finally:
        tmp.unlink(missing_ok=True)


def _check_config_missing() -> None:
    """Ensure ``config.load`` raises when the file is absent."""

    try:
        load("_no_such_config.yml")
    except FileNotFoundError:
        pass
    else:  # pragma: no cover - should not happen
        raise SystemExit("config.load missing-file check failed")


def _ensure_periods_placeholder(base: Path, *, message: str) -> None:
    """Populate exported period files with a placeholder row if empty."""

    csv_path = base.with_name(f"{base.stem}_periods.csv")
    json_path = base.with_name(f"{base.stem}_periods.json")
    txt_path = base.with_name(f"{base.stem}_periods.txt")
    placeholder = {"period": "N/A", "note": message}

    if csv_path.exists():
        try:
            df = pd.read_csv(csv_path)
        except pd.errors.EmptyDataError:
            df = pd.DataFrame()
        if df.empty:
            pd.DataFrame([placeholder]).to_csv(csv_path, index=False)

    if json_path.exists():
        try:
            data = json.loads(json_path.read_text())
        except json.JSONDecodeError:
            data = None
        if data is None or (isinstance(data, list) and len(data) == 0):
            json_path.write_text(json.dumps([placeholder], indent=2))

    if txt_path.exists() and not txt_path.read_text().strip():
        safe_msg = message.replace("\n", " ")
        txt_path.write_text(f"period,note\nN/A,{safe_msg}\n")


def _check_empty_export_helpers() -> None:
    """Ensure export helpers cope with empty result lists."""

    empty: list[dict[str, object]] = []
    pf_frames = export.period_frames_from_results(empty)
    if pf_frames:
        raise SystemExit("period_frames_from_results non-empty for empty input")
    wb_frames = export.workbook_frames_from_results(empty)
    if wb_frames:
        raise SystemExit("workbook_frames_from_results non-empty for empty input")
    path_phase1 = Path("demo/exports/empty_phase1.xlsx")
    export.export_phase1_workbook(empty, str(path_phase1))
    if not path_phase1.exists():
        raise SystemExit("export_phase1_workbook empty failed")
    path_phase1_multi = Path("demo/exports/empty_phase1_multi")
    export.export_phase1_multi_metrics(
        empty,
        str(path_phase1_multi),
        formats=["xlsx", "csv", "json", "txt"],
        include_metrics=True,
    )
    if not path_phase1_multi.with_suffix(".xlsx").exists():
        raise SystemExit("export_phase1_multi_metrics empty Excel missing")
    periods_stub = path_phase1_multi.with_name(f"{path_phase1_multi.stem}_periods.csv")
    if not periods_stub.exists():
        raise SystemExit("export_phase1_multi_metrics empty CSV missing")
    _ensure_periods_placeholder(
        path_phase1_multi,
        message="No multi-period results were generated for the empty demo input.",
    )
    path_multi = Path("demo/exports/empty_multi")
    export.export_multi_period_metrics(
        empty,
        str(path_multi),
        formats=["xlsx", "csv", "json", "txt"],
        include_metrics=True,
    )
    if not path_multi.with_suffix(".xlsx").exists():
        raise SystemExit("export_multi_period_metrics empty Excel missing")
    if not path_multi.with_name(f"{path_multi.stem}_periods.csv").exists():
        raise SystemExit("export_multi_period_metrics empty CSV missing")
    _ensure_periods_placeholder(
        path_multi,
        message="Placeholder row written because no periods were available.",
    )


def _check_export_misc() -> None:
    """Exercise helper utilities in :mod:`export`."""
    tmp = Path("demo/exports/tmp/subdir/test.csv")
    if tmp.parent.exists():
        shutil.rmtree(tmp.parent)
    export._ensure_dir(tmp)
    if not tmp.parent.exists():
        raise SystemExit("_ensure_dir failed")
    df = pd.DataFrame({"A": [1, 2]})
    out = export._apply_format(df, lambda d: d + 1)
    if out.iloc[0, 0] != 2:
        raise SystemExit("_apply_format failed")


def _check_export_content() -> None:
    """Ensure text/CSV/JSON exporters actually write data."""

    prefix = Path("demo/exports/content_test")
    df = pd.DataFrame({"A": [1, 2], "B": [3, 4]})
    export.export_data({"tbl": df}, str(prefix), formats=["csv", "json", "txt"])
    for ext in ("csv", "json", "txt"):
        path = prefix.with_name(f"{prefix.stem}_tbl.{ext}")
        if not path.exists():
            raise SystemExit(f"{ext} export missing")
        data = path.read_text().strip()
        if "A" not in data:
            raise SystemExit(f"{ext} export content corrupt")


def _check_run_full_outputs(cfg: Config) -> None:
    """Validate keys returned by ``pipeline.run_full``."""

    res = pipeline.run_full(cfg)
    required = {
        "selected_funds",
        "score_frame",
        "benchmark_ir",
        "in_sample_scaled",
    }
    missing = required - set(res)
    if missing:
        raise SystemExit(f"run_full missing keys: {missing}")


# Execute additional error handling checks
_check_export_errors()
_check_config_errors()
_check_config_missing()
_check_empty_export_helpers()
_check_export_misc()
_check_export_content()
_check_run_full_outputs(cfg)
_check_run_analysis_errors(cfg)


def _check_package_exports() -> None:
    """Validate ``trend_analysis.__all__`` exposes key modules."""
    expected = {
        "metrics",
        "config",
        "data",
        "pipeline",
        "export",
        "selector",
        "weighting",
        "run_multi_analysis",
    }
    missing = expected - set(ta.__all__)
    if missing:
        raise SystemExit(f"Package __all__ missing: {missing}")
    for name in expected:
        getattr(ta, name)


_check_package_exports()


def _check_module_exports() -> None:
    """Ensure module-level ``__all__`` lists are intact."""

    expected_map = {
        "config": {
            "PresetConfig",
            "ColumnMapping",
            "ConfigurationState",
            "load_preset",
            "list_available_presets",
            "load",
            "load_config",
            "Config",
            "ConfigType",
            "DEFAULTS",
        },
        "data": {"load_csv", "identify_risk_free_fund", "ensure_datetime"},
        "export": {
            "FORMATTERS_EXCEL",
            "register_formatter_excel",
            "reset_formatters_excel",
            "make_summary_formatter",
            "combined_summary_frame",
            "make_period_formatter",
            "format_summary_text",
            "export_to_excel",
            "export_to_csv",
            "export_to_json",
            "export_to_txt",
            "export_data",
            "metrics_from_result",
            "combined_summary_result",
            "summary_frame_from_result",
            "phase1_workbook_data",
            "period_frames_from_results",
            "workbook_frames_from_results",
            "flat_frames_from_results",
            "export_phase1_workbook",
            "export_phase1_multi_metrics",
            "export_multi_period_metrics",
            "export_bundle",
        },
        "weighting": {
            "BaseWeighting",
            "EqualWeight",
            "ScorePropSimple",
            "ScorePropBayesian",
            "AdaptiveBayesWeighting",
        },
        "pipeline": {
            "Stats",
            "calc_portfolio_returns",
            "single_period_run",
            "run_analysis",
            "run",
            "run_full",
        },
        "multi_period": {"run", "Portfolio", "run_schedule"},
        "gui": {
            "launch",
            "load_state",
            "save_state",
            "reset_weight_state",
            "build_config_dict",
            "build_config_from_store",
            "register_plugin",
            "iter_plugins",
            "discover_plugins",
            "ParamStore",
            "debounce",
            "list_builtin_cfgs",
        },
        "core.rank_selection": {
            "FundSelectionConfig",
            "RiskStatsConfig",
            "register_metric",
            "METRIC_REGISTRY",
            "WindowMetricBundle",
            "make_window_key",
            "get_window_metric_bundle",
            "reset_selector_cache",
            "selector_cache_hits",
            "selector_cache_misses",
            "blended_score",
            "compute_metric_series_with_cache",
            "rank_select_funds",
            "selector_cache_stats",
            "clear_window_metric_cache",
            "select_funds",
            "build_ui",
            "canonical_metric_list",
        },
    }

    for name, expected in expected_map.items():
        if "." in name:
            module = importlib.import_module(f"trend_analysis.{name}")
        else:
            module = getattr(ta, name)
        actual = set(getattr(module, "__all__", []))
        if actual != expected:
            raise SystemExit(f"{name} __all__ mismatch")


_check_module_exports()


def _check_cli_help() -> None:
    """Ensure the CLI entry points print help and exit cleanly."""
    env = os.environ.copy()
    env["PYTHONPATH"] = f"{ROOT / 'src'}:{env.get('PYTHONPATH', '')}"
    env.setdefault("TREND_MODEL_SITE_CUSTOMIZE", "1")
    subprocess.run(
        [sys.executable, "-m", "trend_analysis.run_analysis", "--help"],
        check=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        shell=False,
        env=env,
    )
    subprocess.run(
        [sys.executable, "-m", "trend_analysis.run_multi_analysis", "--help"],
        check=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        shell=False,
        env=env,
    )
    subprocess.run(
        [sys.executable, "-m", "trend_analysis.cli", "--help"],
        check=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        shell=False,
        env=env,
    )


_check_cli_help()

# Verify top-level package exports
pkg_cfg = ta.config.load("config/demo.yml")
pkg_df = ta.load_csv(pkg_cfg.data["csv_path"])
if ta.identify_risk_free_fund(pd.DataFrame(pkg_df)) is None:
    raise SystemExit("Package export check failed")
ta.reset_formatters_excel()


@ta.register_formatter_excel("pkg")  # type: ignore[misc]
def _pkg_fmt(
    ws: Any,
    _wb: Any,
) -> None:
    ws.write(0, 0, "pkg")


pkg_path = Path("demo/exports/pkg.xlsx")
ta.export_to_excel({"pkg": pd.DataFrame({"A": [1]})}, str(pkg_path))
if not pkg_path.exists():
    raise SystemExit("Package export_to_excel failed")

# Exercise package-level exporter helpers
pkg_csv = pkg_path.with_suffix(".csv")
ta.export_to_csv({"pkg": pd.DataFrame({"A": [1]})}, str(pkg_csv))
expected_csv = pkg_csv.with_name(f"{pkg_csv.stem}_pkg.csv")
if not expected_csv.exists():
    raise SystemExit("Package export_to_csv failed")
pkg_json = pkg_path.with_suffix(".json")
ta.export_to_json({"pkg": pd.DataFrame({"A": [1]})}, str(pkg_json))
expected_json = pkg_json.with_name(f"{pkg_json.stem}_pkg.json")
if not expected_json.exists():
    raise SystemExit("Package export_to_json failed")
pkg_data = pkg_path.with_name("pkg_data")
ta.export_data({"pkg": pd.DataFrame({"A": [1]})}, str(pkg_data), formats=["txt"])
expected_txt = pkg_data.with_name(f"{pkg_data.stem}_pkg.txt")
if not expected_txt.exists():
    raise SystemExit("Package export_data failed")
pkg_phase1 = pkg_path.with_name("pkg_phase1")
ta.export_phase1_multi_metrics(results, str(pkg_phase1), formats=["csv"])
if not pkg_phase1.with_name(f"{pkg_phase1.stem}_periods.csv").exists():
    raise SystemExit("Package export_phase1_multi_metrics failed")
pkg_multi = pkg_path.with_name("pkg_multi")
ta.export_multi_period_metrics(results, str(pkg_multi), formats=["json"])
if not pkg_multi.with_name(f"{pkg_multi.stem}_periods.json").exists():
    raise SystemExit("Package export_multi_period_metrics failed")
summary_pkg = ta.combined_summary_result(results)
if ta.metrics_from_result(summary_pkg).empty:
    raise SystemExit("Package metrics_from_result failed")

# run_analysis.main directly
if run_analysis.main(["-c", "config/demo.yml"]) != 0:
    raise SystemExit("run_analysis.main failed")
if run_analysis.main(["-c", "config/demo.yml", "--detailed"]) != 0:
    raise SystemExit("run_analysis.main detailed failed")

# Run the multi-period CLI using a temporary config file
cli_cfg = Path("demo/exports/mp_cli_cfg.yml")
cli_out = Path("demo/exports/mp_cli")
data = cfg.model_dump()
data.setdefault("export", {})["directory"] = str(cli_out)
data["export"]["formats"] = ["csv"]
with cli_cfg.open("w", encoding="utf-8") as fh:
    yaml.safe_dump(data, fh)
rc = run_multi_analysis.main(["-c", str(cli_cfg)])
if rc != 0:
    raise SystemExit("run_multi_analysis CLI failed")
if not list(cli_out.glob("*.csv")):
    raise SystemExit("run_multi_analysis CLI produced no output")
rc = run_multi_analysis.main(["-c", str(cli_cfg), "--detailed"])
if rc != 0:
    raise SystemExit("run_multi_analysis CLI detailed failed")
cli_cfg.unlink()


print("Multi-period demo checks passed")

# Run the CLI entry point in both modes to verify it behaves correctly
_env = os.environ.copy()
_env["PYTHONPATH"] = f"{ROOT / 'src'}:{_env.get('PYTHONPATH', '')}"
_env.setdefault("TREND_MODEL_SITE_CUSTOMIZE", "1")
subprocess.run(
    [
        sys.executable,
        "-m",
        "trend_analysis.run_analysis",
        "-c",
        "config/demo.yml",
    ],
    check=True,
    shell=False,
    env=_env,
)
subprocess.run(
    [
        sys.executable,
        "-m",
        "trend_analysis.run_analysis",
        "-c",
        "config/demo.yml",
        "--detailed",
    ],
    check=True,
    shell=False,
    env=_env,
)
subprocess.run(
    [
        sys.executable,
        "-m",
        "trend_analysis.run_multi_analysis",
        "-c",
        "config/demo.yml",
        "--detailed",
    ],
    check=True,
    shell=False,
    env=_env,
)

subprocess.run(
    ["scripts/trend-model", "--check", "run", "-c", "config/demo.yml"],
    check=True,
    shell=False,
)

env = os.environ.copy()
env["TREND_CFG"] = "config/demo.yml"
env.setdefault("TREND_MODEL_SITE_CUSTOMIZE", "1")
subprocess.run(
    ["scripts/trend-model", "--check", "run"],
    check=True,
    env=env,
    shell=False,
)

# Run the validation helpers to ensure they remain functional
quick_check = Path(__file__).resolve().with_name("quick_check.sh")
subprocess.run(["bash", str(quick_check)], check=True, shell=False)

# Execute the full test suite to cover the entire code base
run_tests = Path(__file__).resolve().with_name("run_tests.sh")
result = subprocess.run([str(run_tests)], shell=False)
if result.returncode != 0:
    raise SystemExit(f"{run_tests} failed with exit code {result.returncode}")
