"""Export helpers for trend analysis results."""

from __future__ import annotations

import inspect
import math
from collections import OrderedDict
from pathlib import Path
from typing import Any, Callable, Iterable, Mapping, Sequence, cast

try:  # Optional openpyxl for richer typing; not required at runtime.
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.worksheet import Worksheet
except Exception:  # pragma: no cover - openpyxl not installed
    Worksheet = Any  # fallback alias used when openpyxl is absent
    get_column_letter = None

import numpy as np
import pandas as pd

from . import bundle as bundle  # noqa: F401  # re-exported module for tests/compat
from .bundle import export_bundle

Formatter = Callable[[pd.DataFrame], pd.DataFrame]


FORMATTERS_EXCEL: dict[str, Callable[[Any, Any], None]] = {}

_OPENPYXL_COLOR_MAP = {
    "red": "FFFF0000",
}


def _normalise_color(value: Any) -> str | None:
    """Return an ARGB hex colour string understood by openpyxl."""

    if not isinstance(value, str):
        return None
    stripped = value.strip()
    if not stripped:
        return None
    mapped = _OPENPYXL_COLOR_MAP.get(stripped.lower())
    if mapped:
        return mapped
    if stripped.startswith("#"):
        stripped = stripped[1:]
    length = len(stripped)
    if length == 6:
        return f"FF{stripped.upper()}"
    if length == 8:
        return stripped.upper()
    return None


def _is_openpyxl_book(book: Any) -> bool:
    """Return ``True`` when ``book`` looks like an openpyxl workbook."""

    module = getattr(book.__class__, "__module__", "")
    return module.startswith("openpyxl")


def _maybe_remove_openpyxl_default_sheet(book: Any) -> str | None:
    """Drop the empty default sheet created by openpyxl workbooks.

    Returns the title of the removed sheet when a removal occurs so callers can
    keep auxiliary mappings (e.g. :attr:`pandas.ExcelWriter.sheets`) in sync.
    """

    try:
        worksheets = list(getattr(book, "worksheets", []))
    except Exception:  # pragma: no cover - defensive guard
        return None
    if len(worksheets) != 1:
        return None
    ws = worksheets[0]
    title = getattr(ws, "title", "")
    if title.lower() != "sheet":
        return None
    try:
        cell = ws.cell(row=1, column=1)
    except Exception:  # pragma: no cover - defensive guard
        return None
    if getattr(cell, "value", None) is None:
        try:
            book.remove(ws)
            return title
        except Exception:  # pragma: no cover - defensive guard
            return None
    return None


class _OpenpyxlWorksheetProxy:
    """Adapter exposing the subset of ``xlsxwriter`` APIs used by
    formatters."""

    def __init__(self, ws: Any):
        self._ws = ws

    @property
    def name(self) -> str:
        return getattr(self._ws, "title", "")

    def write(self, row: int, col: int, value: Any, fmt: Any | None = None) -> None:
        cell = self._ws.cell(row=row + 1, column=col + 1)
        cell.value = value
        if fmt and isinstance(fmt, dict):
            self._apply_format(cell, fmt)

    def write_row(
        self, row: int, col: int, data: Iterable[Any], fmt: Any | None = None
    ) -> None:
        for offset, value in enumerate(data):
            self.write(row, col + offset, value, fmt)

    def set_column(self, first_col: int, last_col: int, width: float) -> None:
        if get_column_letter is None:
            return
        for col in range(first_col, last_col + 1):
            letter = get_column_letter(col + 1)
            dim = self._ws.column_dimensions[letter]
            dim.width = width

    def freeze_panes(self, row: int, col: int) -> None:
        self._ws.freeze_panes = self._ws.cell(row=row + 1, column=col + 1)

    def autofilter(self, fr: int, fc: int, lr: int, lc: int) -> None:
        if get_column_letter is None:
            return
        start_col = get_column_letter(fc + 1)
        end_col = get_column_letter(lc + 1)
        self._ws.auto_filter.ref = f"{start_col}{fr + 1}:{end_col}{lr + 1}"

    def _apply_format(self, cell: Any, fmt: Mapping[str, Any]) -> None:
        if "num_format" in fmt:
            num_format = fmt["num_format"]
            if hasattr(cell, "number_format"):
                cell.number_format = num_format
        if "font_color" in fmt:
            colour_hex = _normalise_color(fmt["font_color"])
            if colour_hex:
                font = getattr(cell, "font", None)
                if font is not None and hasattr(font, "copy"):
                    cell.font = font.copy(color=colour_hex)


class _OpenpyxlWorkbookProxy:
    """Expose workbook helpers expected by the registered formatters."""

    def __init__(self, writer: Any):
        self._writer = writer

    @property
    def _book(self) -> Any:
        return self._writer.book

    def add_format(self, spec: Mapping[str, Any]) -> Mapping[str, Any]:
        # Formatting support is best-effort under openpyxl; return the spec so
        # callers can still pass it back to ``write``.
        return dict(spec)

    def add_worksheet(self, name: str) -> _OpenpyxlWorksheetProxy:
        book = self._book
        removed = _maybe_remove_openpyxl_default_sheet(book)
        if removed:
            self._writer.sheets.pop(removed, None)
        ws = book.create_sheet(title=name)
        return _OpenpyxlWorksheetProxy(ws)

    def __getattr__(self, name: str) -> Any:  # pragma: no cover - simple proxy
        return getattr(self._book, name)


class _OpenpyxlWorksheetAdapter:
    """Lightweight adapter exposing a subset of the xlsxwriter worksheet
    API."""

    __slots__ = ("_ws",)

    def __init__(self, worksheet: Any) -> None:
        self._ws = worksheet

    @property
    def native(self) -> Any:
        return self._ws

    def write(
        self, row: int, col: int, value: object, fmt: object | None = None
    ) -> None:  # noqa: ARG002
        # The `fmt` parameter is ignored because openpyxl's formatting model is
        # different from xlsxwriter's, and this adapter does not support cell formatting.
        self._ws.cell(row=row + 1, column=col + 1, value=value)

    def write_row(
        self,
        row: int,
        col: int,
        data: Iterable[object],
        fmt: object | None = None,  # noqa: ARG002
    ) -> None:
        for offset, value in enumerate(data):
            self.write(row, col + offset, value)

    def set_column(self, first_col: int, last_col: int, width: float) -> None:
        from openpyxl.utils import get_column_letter

        for idx in range(first_col, last_col + 1):
            letter = get_column_letter(idx + 1)
            self._ws.column_dimensions[letter].width = width

    def freeze_panes(self, row: int, col: int) -> None:
        self._ws.freeze_panes = self._ws.cell(row=row + 1, column=col + 1)

    def autofilter(self, fr: int, fc: int, lr: int, lc: int) -> None:
        from openpyxl.utils import get_column_letter

        start = f"{get_column_letter(fc + 1)}{fr + 1}"
        end = f"{get_column_letter(lc + 1)}{lr + 1}"
        self._ws.auto_filter.ref = f"{start}:{end}"


class _OpenpyxlWorkbookAdapter:
    """Adapter that exposes minimal workbook hooks expected by formatters."""

    __slots__ = ("_wb",)

    def __init__(self, workbook: Any) -> None:
        self._wb = workbook
        self._prune_default_sheet()

    def _prune_default_sheet(self) -> None:
        sheets = getattr(self._wb, "worksheets", [])
        if len(sheets) == 1:
            sheet = sheets[0]
            title = getattr(sheet, "title", "")
            value = sheet.cell(1, 1).value if hasattr(sheet, "cell") else None
            if title == "Sheet" and value in (None, "") and hasattr(self._wb, "remove"):
                self._wb.remove(sheet)

    def add_worksheet(self, name: str) -> _OpenpyxlWorksheetAdapter:
        ws = self._wb.create_sheet(title=name)
        if getattr(ws, "title", name) != name:
            ws.title = name
        return _OpenpyxlWorksheetAdapter(ws)

    def add_format(self, spec: Mapping[str, Any] | None) -> Mapping[str, Any]:
        return dict(spec or {})

    def rename_last_sheet(self, name: str) -> None:
        sheets = getattr(self._wb, "worksheets", [])
        if sheets:
            sheets[-1].title = name

    def __getattr__(self, attr: str) -> Any:
        return getattr(self._wb, attr)


def register_formatter_excel(
    category: str,
) -> Callable[[Callable[[Any, Any], None]], Callable[[Any, Any], None]]:
    """Register an Excel formatter under ``category``."""

    def decorator(fn: Callable[[Any, Any], None]) -> Callable[[Any, Any], None]:
        FORMATTERS_EXCEL[category] = fn
        return fn

    return decorator


def reset_formatters_excel() -> None:
    """Clear all registered Excel formatters."""
    FORMATTERS_EXCEL.clear()


def _format_frequency_policy_line(res: Mapping[str, Any]) -> str:
    freq = cast(Mapping[str, Any], res.get("input_frequency", {}))
    if not freq:
        freq = cast(Mapping[str, Any], res.get("input_frequency_details", {}))
    if not freq:
        freq = cast(Mapping[str, Any], res.get("preprocessing", {})).get(
            "input_frequency_details", {}
        )
    if not freq:
        metadata = res.get("metadata")
        if isinstance(metadata, Mapping):
            freq = cast(Mapping[str, Any], metadata.get("frequency", {}))

    policy = cast(Mapping[str, Any], res.get("missing_data_policy", {}))
    if not policy:
        policy = cast(Mapping[str, Any], res.get("preprocessing", {})).get(
            "missing_data_policy", {}
        )
    if not policy:
        metadata = res.get("metadata")
        if isinstance(metadata, Mapping):
            policy = cast(Mapping[str, Any], metadata.get("missing_data", {}))

    freq_label = cast(str | None, freq.get("label"))
    target_label = cast(str | None, freq.get("target_label")) or freq_label
    resampled = bool(freq.get("resampled"))
    if freq_label:
        if resampled and target_label and target_label != freq_label:
            freq_part = f"{freq_label} → {target_label}"
        else:
            freq_part = freq_label
    else:
        freq_part = "Unknown"

    policy_name = str(policy.get("policy", "drop")).lower()
    policy_labels = {
        "drop": "Drop",
        "ffill": "Forward-fill",
        "zero": "Zero-fill",
    }
    policy_part = policy_labels.get(policy_name, policy_name.title())
    extras: list[str] = []
    limit = policy.get("limit")
    if policy_name == "ffill" and limit is not None:
        extras.append(f"limit={limit}")
    total_filled = policy.get("total_filled")
    filled_int = 0
    if total_filled is not None:
        try:
            filled_int = int(total_filled)
        except (TypeError, ValueError):
            filled_int = 0
    if filled_int:
        extras.append(f"filled {filled_int} cell{'s' if filled_int != 1 else ''}")
    dropped_assets = policy.get("dropped_assets")
    dropped_count = len(dropped_assets) if isinstance(dropped_assets, list) else 0
    if dropped_count:
        extras.append(
            f"dropped {dropped_count} asset{'s' if dropped_count != 1 else ''}"
        )
    if extras:
        policy_part = f"{policy_part} ({', '.join(extras)})"
    return f"Frequency: {freq_part}; NA policy: {policy_part}"


def _format_metadata_entries(meta: Mapping[str, Any]) -> list[str]:
    lines: list[str] = []
    universe = meta.get("universe")
    if isinstance(universe, Mapping):
        members = universe.get("members")
        members_list = list(members) if isinstance(members, Sequence) else []
        total = cast(int | None, universe.get("count")) or len(members_list)
        selected = universe.get("selected")
        selected_list = list(selected) if isinstance(selected, Sequence) else []
        selected_count = cast(int | None, universe.get("selected_count")) or len(
            selected_list
        )
        line = "Universe: "
        line += f"{total} assets" if total else "N/A"
        if selected_count:
            line += f"; selected {selected_count}"
        preview_source = selected_list or members_list
        if preview_source:
            preview = ", ".join(str(val) for val in preview_source[:5])
            if preview:
                suffix = "…" if len(preview_source) > 5 else ""
                line += f" ({preview}{suffix})"
        lines.append(line)

    lookbacks = meta.get("lookbacks")
    if isinstance(lookbacks, Mapping):
        in_sample = cast(Mapping[str, Any], lookbacks.get("in_sample", {}))
        out_sample = cast(Mapping[str, Any], lookbacks.get("out_sample", {}))
        if in_sample or out_sample:
            in_part = (
                f"In {in_sample.get('start')} → {in_sample.get('end')}"
                if in_sample.get("start") or in_sample.get("end")
                else "In (unspecified)"
            )
            out_part = (
                f"Out {out_sample.get('start')} → {out_sample.get('end')}"
                if out_sample.get("start") or out_sample.get("end")
                else "Out (unspecified)"
            )
            lines.append(f"Lookbacks: {in_part}; {out_part}")

    costs = meta.get("costs")
    if isinstance(costs, Mapping) and costs:
        formatted: list[str] = []
        for key, value in costs.items():
            try:
                number = float(value)
            except (TypeError, ValueError):
                continue
            if math.isnan(number):
                continue
            label = str(key)
            if abs(number) <= 1:
                formatted.append(f"{label}={number:.2%}")
            else:
                formatted.append(f"{label}={number:.4f}")
        if formatted:
            lines.append("Costs: " + ", ".join(formatted))

    code_version = meta.get("code_version")
    fingerprint = meta.get("fingerprint")
    parts: list[str] = []
    if code_version:
        parts.append(f"Code version: {code_version}")
    if fingerprint:
        parts.append(f"Fingerprint: {fingerprint}")
    if parts:
        lines.append("; ".join(parts))
    return lines


def _metadata_summary_lines(res: Mapping[str, Any]) -> list[str]:
    segments: list[str] = []
    summary_text = cast(
        str,
        res.get("preprocessing_summary")
        or cast(Mapping[str, Any], res.get("preprocessing", {})).get("summary"),
    )
    if summary_text:
        segments.append(summary_text)
    freq_line = _format_frequency_policy_line(res)
    if freq_line:
        segments.append(freq_line)
    metadata = res.get("metadata")
    if isinstance(metadata, Mapping):
        segments.extend(_format_metadata_entries(metadata))
    if not segments:
        segments.append("Frequency: Unknown; NA policy: (not specified)")
    # The legacy workbook layout expects a single metadata row before the
    # tabular header so downstream consumers can ``skiprows=4`` reliably.
    # Combine all entries into one readable line separated by " | " to
    # preserve the new metadata while keeping the row budget constant.
    return [" | ".join(segments)]


def _build_summary_formatter(
    res: Mapping[str, Any],
    in_start: str,
    in_end: str,
    out_start: str,
    out_end: str,
) -> Callable[[Any, Any], None]:
    """Return a formatter function for a summary sheet."""

    def fmt_summary(ws: Any, wb: Any) -> None:
        bold = wb.add_format({"bold": True})
        num2 = wb.add_format({"num_format": "0.00"})
        pct2 = wb.add_format({"num_format": "0.00%"})
        pct2_red = wb.add_format({"num_format": "0.00%", "font_color": "red"})

        def safe(v: float | str | None) -> str | float:
            if v is None:
                return ""
            if isinstance(v, float):
                if math.isnan(v):
                    return ""
                return v
            if pd.isna(v):
                return ""
            return v if isinstance(v, (int, float, str)) else ""

        def to_tuple(obj: Any) -> tuple[float, float, float, float, float, float]:
            if isinstance(obj, tuple):
                return cast(tuple[float, float, float, float, float, float], obj)
            return (
                cast(float, obj.cagr),
                cast(float, obj.vol),
                cast(float, obj.sharpe),
                cast(float, obj.sortino),
                cast(float, obj.information_ratio),
                cast(float, obj.max_drawdown),
            )

        # Metrics list in raw units expected by cell formatters.
        # CAGR/Vol/MaxDD are fractions (0..1) to be rendered with % format.
        def metrics_list(t: Any) -> list[float]:
            tup = to_tuple(t)
            return [tup[0], tup[1], tup[2], tup[3], tup[4], tup[5]]

        ws.write_row(0, 0, ["Vol-Adj Trend Analysis"], bold)
        ws.write_row(1, 0, [f"In:  {in_start} → {in_end}"], bold)
        ws.write_row(2, 0, [f"Out: {out_start} → {out_end}"], bold)
        meta_lines = _metadata_summary_lines(res)
        start_row = 3
        for offset, line in enumerate(meta_lines):
            ws.write_row(start_row + offset, 0, [line])
        bench_labels = list(res.get("benchmark_ir", {}))
        headers = [
            "Name",
            "Weight",
            "IS CAGR",
            "IS Vol",
            "IS Sharpe",
            "IS Sortino",
            "IS IR",
            "IS MaxDD",
            "OS CAGR",
            "OS Vol",
            "OS Sharpe",
            "OS Sortino",
            "OS IR",
        ]
        headers.extend([f"OS IR {b}" for b in bench_labels])
        headers.append("OS MaxDD")
        header_row = start_row + len(meta_lines)
        ws.write_row(header_row, 0, headers, bold)
        for idx, h in enumerate(headers):
            ws.set_column(idx, idx, len(h) + 2)
        ws.freeze_panes(header_row + 1, 0)
        numeric_fmts: list[Any] = []
        for h in headers[2:]:
            if "MaxDD" in h:
                numeric_fmts.append(pct2_red)
            elif "CAGR" in h or "Vol" in h:
                numeric_fmts.append(pct2)
            else:
                numeric_fmts.append(num2)

        row = header_row + 1
        for label, ins, outs in [
            ("Equal Weight", res["in_ew_stats"], res["out_ew_stats"]),
            ("User Weight", res["in_user_stats"], res["out_user_stats"]),
        ]:
            ws.write(row, 0, label, bold)
            ws.write(row, 1, safe(""))
            ins_vals = metrics_list(ins)
            outs_vals = metrics_list(outs)
            # Defer OS MaxDD to the final column after benchmark IRs
            os_maxdd = outs_vals[-1]
            vals = ins_vals + outs_vals[:-1]
            extra = [
                res.get("benchmark_ir", {})
                .get(b, {})
                .get("equal_weight" if label == "Equal Weight" else "user_weight", "")
                for b in bench_labels
            ]
            fmts = numeric_fmts
            vals.extend(extra)
            vals.append(os_maxdd)
            for col, (v, fmt) in enumerate(zip(vals, fmts), start=2):
                ws.write(row, col, safe(v), fmt)
            row += 1

        # Start fund rows immediately after the aggregate rows (no spacer),
        # so the first fund appears on row 8 (1-based indexing).
        for fund, stat_in in res["in_sample_stats"].items():
            stat_out = res["out_sample_stats"][fund]
            ws.write(row, 0, fund, bold)
            wt = res["fund_weights"][fund]
            # Write weights as fractions with percent formatting
            ws.write(row, 1, safe(wt), pct2)
            ins_vals = metrics_list(stat_in)
            outs_vals = metrics_list(stat_out)
            os_maxdd = outs_vals[-1]
            vals = ins_vals + outs_vals[:-1]
            extra = [
                res.get("benchmark_ir", {}).get(b, {}).get(fund, "")
                for b in bench_labels
            ]
            fmts = numeric_fmts
            vals.extend(extra)
            vals.append(os_maxdd)
            for col, (v, fmt) in enumerate(zip(vals, fmts), start=2):
                ws.write(row, col, safe(v), fmt)
            row += 1

        ws.autofilter(header_row, 0, row - 1, len(headers) - 1)

        # Optional: append a Manager Changes section after the main table.
        changes = cast(list[Mapping[str, Any]] | None, res.get("manager_changes"))
        if changes:
            row += 2
            ws.write_row(row, 0, ["Manager Changes"], bold)
            row += 1
            # Determine available columns from the first record and keep a stable order
            # Prefer a canonical subset if present.
            preferred = ["Period", "action", "manager", "firm", "reason", "detail"]
            keys = list({k for rec in changes for k in rec.keys()})
            ordered = [k for k in preferred if k in keys] + [
                k for k in keys if k not in preferred
            ]
            ws.write_row(
                row, 0, [k.capitalize() if k != "Period" else k for k in ordered], bold
            )
            # Set reasonable widths
            for idx, k in enumerate(ordered):
                width = 12 if k in {"Period", "action", "firm"} else 24
                ws.set_column(idx, idx, width)
            row += 1
            for rec in changes:
                vals = [rec.get(k, "") for k in ordered]
                for col, v in enumerate(vals):
                    ws.write(row, col, v)
                row += 1

        # Optional: append a Manager Participation & Contribution section
        contrib = res.get("manager_contrib")
        if contrib is not None:
            # Accept either a DataFrame-like (records) or a list of dicts
            # Normalize to a list of dict rows with expected keys
            rows: list[dict[str, Any]]
            if isinstance(contrib, pd.DataFrame):
                df2 = contrib.copy()
                df2.columns = [str(c) for c in df2.columns]
                rows = cast(list[dict[str, Any]], df2.to_dict(orient="records"))
            else:
                rows = [dict(r) for r in cast(list[Mapping[str, Any]], contrib)]
            if rows:
                row += 2
                ws.write_row(row, 0, ["Manager Participation & Contribution"], bold)
                row += 1
                headers = ["Manager", "Years", "OOS CAGR", "Contribution Share"]
                ws.write_row(row, 0, headers, bold)
                # Set column widths and number formats
                ws.set_column(0, 0, 30)  # Manager
                ws.set_column(1, 1, 10)  # Years
                ws.set_column(2, 3, 18)  # Rates
                row += 1
                for rec in rows:
                    ws.write(row, 0, rec.get("Manager", ""))
                    ws.write(row, 1, rec.get("Years", ""), num2)
                    ws.write(row, 2, rec.get("OOS CAGR", ""), pct2)
                    ws.write(row, 3, rec.get("Contribution Share", ""), pct2)
                    row += 1

    return fmt_summary


def make_summary_formatter(
    res: Mapping[str, Any],
    in_start: str,
    in_end: str,
    out_start: str,
    out_end: str,
) -> Callable[[Any, Any], None]:
    """Return and register a formatter for the ``summary`` sheet."""

    fmt = _build_summary_formatter(res, in_start, in_end, out_start, out_end)
    return register_formatter_excel("summary")(fmt)


def make_period_formatter(
    sheet: str,
    res: Mapping[str, Any],
    in_start: str,
    in_end: str,
    out_start: str,
    out_end: str,
) -> Callable[[Any, Any], None]:
    """Return and register a formatter for a per-period sheet."""

    fmt = _build_summary_formatter(res, in_start, in_end, out_start, out_end)
    return register_formatter_excel(sheet)(fmt)


def format_summary_text(
    res: Mapping[str, Any],
    in_start: str,
    in_end: str,
    out_start: str,
    out_end: str,
) -> str:
    """Return a plain-text summary table similar to the Excel output."""

    def safe(val: float | int | str | None) -> str:
        if val is None:
            return ""
        if isinstance(val, float):
            if math.isnan(val):
                return ""
            return f"{val:.2f}"
        if isinstance(val, int):
            return f"{float(val):.2f}"
        # strings or other objects
        return str(val)

    def to_tuple(obj: Any) -> tuple[float, float, float, float, float, float]:
        if isinstance(obj, tuple):
            return cast(tuple[float, float, float, float, float, float], obj)
        return (
            cast(float, obj.cagr),
            cast(float, obj.vol),
            cast(float, obj.sharpe),
            cast(float, obj.sortino),
            cast(float, obj.information_ratio),
            cast(float, obj.max_drawdown),
        )

    def pct(t: Any) -> list[float]:
        a = to_tuple(t)
        return [a[0] * 100, a[1] * 100, a[2], a[3], a[4], a[5] * 100]

    bench_map = cast(Mapping[str, Mapping[str, float]], res.get("benchmark_ir", {}))
    bench_labels = list(bench_map)
    columns = [
        "Name",
        "Weight",
        "IS CAGR",
        "IS Vol",
        "IS Sharpe",
        "IS Sortino",
        "IS IR",
        "IS MaxDD",
        "OS CAGR",
        "OS Vol",
        "OS Sharpe",
        "OS Sortino",
        "OS IR",
    ]
    columns.extend([f"OS IR {b}" for b in bench_labels])
    columns.append("OS MaxDD")

    rows: list[list[str | float | None]] = []

    for label, ins, outs in [
        ("Equal Weight", res["in_ew_stats"], res["out_ew_stats"]),
        ("User Weight", res["in_user_stats"], res["out_user_stats"]),
    ]:
        vals = pct(ins) + pct(outs)
        extra = [
            res.get("benchmark_ir", {})
            .get(b, {})
            .get(
                "equal_weight" if label == "Equal Weight" else "user_weight",
                float("nan"),
            )
            for b in bench_labels
        ]
        vals.extend(extra)
        rows.append([label, None, *vals])

    rows.append([None] * len(columns))

    for fund, stat_in in res["in_sample_stats"].items():
        stat_out = res["out_sample_stats"][fund]
        weight = res["fund_weights"][fund] * 100
        vals = pct(stat_in) + pct(stat_out)
        extra = [
            res.get("benchmark_ir", {}).get(b, {}).get(fund, float("nan"))
            for b in bench_labels
        ]
        vals.extend(extra)
        rows.append([fund, weight, *vals])

    df = pd.DataFrame(rows, columns=columns)
    df_formatted = df.map(safe)
    header = [
        "Vol-Adj Trend Analysis",
        f"In:  {in_start} → {in_end}",
        f"Out: {out_start} → {out_end}",
    ]
    header.extend(_metadata_summary_lines(res))
    risk_diag = cast(Mapping[str, Any] | None, res.get("risk_diagnostics"))
    if isinstance(risk_diag, Mapping) and risk_diag:
        asset_vol = risk_diag.get("asset_volatility")
        if isinstance(asset_vol, pd.DataFrame) and not asset_vol.empty:
            header.append("")
            header.append("Realised volatility (latest):")
            latest_row = asset_vol.iloc[-1]
            for name, value in latest_row.items():
                header.append(f"  {name}: {float(value):.2%}")
        port_vol = risk_diag.get("portfolio_volatility")
        if isinstance(port_vol, pd.Series) and not port_vol.empty:
            header.append(
                f"Portfolio volatility (latest): {float(port_vol.iloc[-1]):.2%}"
            )
        turnover_value = risk_diag.get("turnover_value")
        if isinstance(turnover_value, (float, int)) and not math.isnan(
            float(turnover_value)
        ):
            header.append(f"Turnover applied: {float(turnover_value):.2%}")
    header.extend(
        [
            "",
            df_formatted.to_string(index=False),
        ]
    )

    regime_summary = res.get("regime_summary")
    regime_table = res.get("performance_by_regime")
    if isinstance(regime_table, pd.DataFrame) and not regime_table.empty:
        header.append("")
        header.append("Performance by regime (CAGR / Sharpe / MaxDD / Hit / Obs):")

        def _fmt_pct(value: Any, digits: int = 1) -> str:
            if pd.isna(value):
                return "N/A"
            return f"{float(value):.{digits}%}"

        def _fmt_ratio(value: Any, digits: int = 2) -> str:
            if pd.isna(value):
                return "N/A"
            return f"{float(value):.{digits}f}"

        def _fmt_count(value: Any) -> str:
            if pd.isna(value):
                return "N/A"
            return f"{float(value):.0f}"

        def _has_metrics(series: pd.Series) -> bool:
            return any(
                pd.notna(series.get(metric))
                for metric in ("CAGR", "Sharpe", "Max Drawdown", "Hit Rate")
            )

        for portfolio in dict.fromkeys(col[0] for col in regime_table.columns):
            header.append(f"  {portfolio}:")
            for _, regime in [
                col for col in regime_table.columns if col[0] == portfolio
            ]:
                series = regime_table[(portfolio, regime)]
                if not _has_metrics(series):
                    header.append(f"    {regime}: insufficient data")
                    continue
                header.append(
                    "    {regime}: {cagr} CAGR, Sharpe {sharpe}, MaxDD {mdd}, "
                    "Hit {hit}, Obs {obs}".format(
                        regime=regime,
                        cagr=_fmt_pct(series.get("CAGR")),
                        sharpe=_fmt_ratio(series.get("Sharpe")),
                        mdd=_fmt_pct(series.get("Max Drawdown")),
                        hit=_fmt_pct(series.get("Hit Rate")),
                        obs=_fmt_count(series.get("Observations")),
                    )
                )

    regime_notes = [str(note).strip() for note in res.get("regime_notes", [])]
    regime_notes = [note for note in regime_notes if note]
    summary_text = regime_summary.strip() if isinstance(regime_summary, str) else ""
    if summary_text:
        header.append("")
        header.append(f"Regime insight: {summary_text}")
    if regime_notes:
        header.append("Regime notes:")
        header.extend(f"  - {note}" for note in regime_notes)
    return "\n".join(header)


def _ensure_dir(path: Path) -> None:
    """Create parent directories for the given path."""
    path.parent.mkdir(parents=True, exist_ok=True)


def _apply_format(df: pd.DataFrame, formatter: Formatter | None) -> pd.DataFrame:
    """Return ``df`` after applying the optional ``formatter``."""
    return formatter(df) if formatter else df


def export_to_excel(
    data: Mapping[str, pd.DataFrame],
    output_path: str,
    formatter: Formatter | None = None,
    default_sheet_formatter: Callable[[Any, Any], None] | None = None,
) -> None:
    """Export dataframes to an Excel workbook.

    Each key in ``data`` becomes a sheet.  After writing a sheet, a
    formatter function is looked up in :data:`FORMATTERS_EXCEL` by sheet name
    and applied.  If absent, ``default_sheet_formatter`` is used if provided.
    The ``formatter`` argument still allows per-frame transformations before
    writing.
    """
    path = Path(output_path)
    _ensure_dir(path)

    df_formatter = formatter
    if formatter and default_sheet_formatter is None:
        params = list(inspect.signature(formatter).parameters)
        if len(params) != 1:
            default_sheet_formatter = cast(Callable[[Any, Any], None], formatter)
            # backward compat
            df_formatter = None

    workbook_adapter: _OpenpyxlWorkbookAdapter | None = None
    try:
        writer = pd.ExcelWriter(path, engine="xlsxwriter")
        supports_custom = True
    except ModuleNotFoundError:
        # ``xlsxwriter`` is an optional dependency.  Fall back to the default
        # engine (typically ``openpyxl``) when it is unavailable so callers can
        # still export workbooks without installing the extra package.
        writer = pd.ExcelWriter(path)
        book_any: Any = writer.book
        supports_custom = all(
            hasattr(book_any, attr) for attr in ("add_worksheet", "add_format")
        )
        if not supports_custom:
            workbook_adapter = _OpenpyxlWorkbookAdapter(book_any)
            supports_custom = True

    with writer:
        workbook_any: Any = getattr(writer, "book", None)
        engine_name = getattr(writer, "engine", None)
        proxy: _OpenpyxlWorkbookProxy | None = None
        supports_sheet_formatters = bool(
            workbook_any and callable(getattr(workbook_any, "add_worksheet", None))
        )
        removed_title: str | None = None
        if not supports_sheet_formatters:
            if workbook_any is not None and _is_openpyxl_book(workbook_any):
                removed_title = _maybe_remove_openpyxl_default_sheet(workbook_any)
                if removed_title:
                    writer.sheets.pop(removed_title, None)
                proxy = _OpenpyxlWorkbookProxy(writer)
                supports_sheet_formatters = True
            elif engine_name == "openpyxl":
                proxy = _OpenpyxlWorkbookProxy(writer)
                supports_sheet_formatters = True
            else:
                proxy = None
        else:
            proxy = None
        if proxy is not None:
            pass
        # Iterate over frames and either let a registered sheet formatter
        # render the entire sheet (preferred), or fall back to writing the
        # DataFrame directly when no formatter is available.
        for sheet, df in data.items():
            fmt = FORMATTERS_EXCEL.get(sheet, default_sheet_formatter)
            if fmt is not None and supports_sheet_formatters:
                if proxy is not None:
                    ws_proxy = proxy.add_worksheet(sheet)
                    writer.sheets[sheet] = ws_proxy._ws
                    fmt(ws_proxy, proxy)
                else:
                    # Create an empty worksheet and delegate full rendering
                    # xlsxwriter workbook object provides add_worksheet; cast for typing
                    workbook_obj = writer.book
                    add_ws = getattr(workbook_obj, "add_worksheet")
                    ws = cast(Worksheet, add_ws(sheet))
                    writer.sheets[sheet] = ws
                    fmt(ws, writer.book)
            else:
                if proxy is not None:
                    removed = _maybe_remove_openpyxl_default_sheet(writer.book)
                    if removed:
                        writer.sheets.pop(removed, None)
                formatted = _apply_format(df, df_formatter)
                formatted.to_excel(writer, sheet_name=sheet, index=False)
                if workbook_adapter is not None:
                    try:
                        workbook_adapter.rename_last_sheet(sheet)
                    except Exception:  # pragma: no cover - defensive best effort
                        pass
                if proxy is not None:
                    try:
                        ws_obj = writer.book[sheet]
                    except KeyError:
                        ws_obj = writer.book.worksheets[-1]
                    current_title = getattr(ws_obj, "title", sheet)
                    if current_title != sheet:
                        try:
                            ws_obj.title = sheet
                        except Exception:  # pragma: no cover - best effort rename
                            pass
                        else:
                            writer.sheets.pop(current_title, None)
                    writer.sheets[sheet] = ws_obj


def export_to_csv(
    data: Mapping[str, pd.DataFrame],
    output_path: str,
    formatter: Formatter | None = None,
) -> None:
    """Export each dataframe to an individual CSV file using ``output_path`` as
    prefix."""
    prefix = Path(output_path)
    _ensure_dir(prefix)
    # Looping over the ``data`` dictionary ensures each frame gets its own file.
    for name, df in data.items():
        formatted = _apply_format(df, formatter)
        formatted.to_csv(
            prefix.with_name(f"{prefix.stem}_{name}.csv"),
            index=False,
            header=True,
        )


def export_to_json(
    data: Mapping[str, pd.DataFrame],
    output_path: str,
    formatter: Formatter | None = None,
) -> None:
    """Export each DataFrame to an individual JSON file using ``output_path``
    as prefix."""
    prefix = Path(output_path)
    _ensure_dir(prefix)
    # Iterate over the mapping so each DataFrame is written to its own JSON file.
    for name, df in data.items():
        formatted = _apply_format(df, formatter)
        formatted.to_json(
            prefix.with_name(f"{prefix.stem}_{name}.json"), orient="records", indent=2
        )


def export_to_txt(
    data: Mapping[str, pd.DataFrame],
    output_path: str,
    formatter: Formatter | None = None,
) -> None:
    """Export each dataframe to a plain text file using ``output_path`` as
    prefix."""
    prefix = Path(output_path)
    _ensure_dir(prefix)
    for name, df in data.items():
        formatted = _apply_format(df, formatter)
        prefix.with_name(f"{prefix.stem}_{name}.txt").write_text(
            formatted.to_string(index=False)
        )


EXPORTERS: dict[
    str, Callable[[Mapping[str, pd.DataFrame], str, Formatter | None], None]
] = {
    "xlsx": export_to_excel,
    # ``excel`` is kept for backward compatibility with older configs/UI
    "excel": export_to_excel,
    "csv": export_to_csv,
    "json": export_to_json,
    "txt": export_to_txt,
}


def execution_metrics_frame(
    results: Iterable[Mapping[str, object]],
) -> pd.DataFrame:
    """Build a per-period execution metrics frame.

    Columns:
      - Period: period label (OOS end month or synthetic label)
      - Turnover: sum of absolute weight changes applied for the period
      - Transaction Cost: linear cost computed from turnover and cost bps (if available)

    Missing values are filled with NaN when a metric isn't present in the period result.
    """
    rows: list[dict[str, Any]] = []
    for idx, res in enumerate(results, start=1):
        period = res.get("period")
        if isinstance(period, (list, tuple)) and len(period) >= 4:
            label = str(period[3])
        else:
            label = f"period_{idx}"
        turnover = cast(float | None, res.get("turnover"))
        tx_cost = cast(float | None, res.get("transaction_cost"))
        rows.append(
            {
                "Period": label,
                "Turnover": float(turnover) if turnover is not None else float("nan"),
                "Transaction Cost": (
                    float(tx_cost) if tx_cost is not None else float("nan")
                ),
            }
        )
    return (
        pd.DataFrame(rows, columns=["Period", "Turnover", "Transaction Cost"])
        if rows
        else pd.DataFrame(columns=["Period", "Turnover", "Transaction Cost"])
    )


def export_execution_metrics(
    results: Iterable[Mapping[str, object]],
    output_path: str,
    *,
    formats: Iterable[str] = ("xlsx",),
) -> None:
    """Export execution metrics (turnover and transaction cost) separately.

    This does not alter the Phase‑1 summary tables or existing exports.

    Produces:
      - Excel: a workbook with a single sheet named ``execution_metrics``
      - CSV/JSON/TXT: files with suffix ``_execution_metrics``
    """
    df = execution_metrics_frame(list(results))
    data = {"execution_metrics": df}
    export_data(data, output_path, formats=formats)


def metrics_from_result(res: Mapping[str, object]) -> pd.DataFrame:
    """Return a metrics DataFrame identical to :func:`pipeline.run` output."""
    from ..pipeline import _Stats  # lazy import to avoid cycle

    stats = cast(Mapping[str, _Stats], res.get("out_sample_stats", {}))
    df = pd.DataFrame({k: vars(v) for k, v in stats.items()}).T
    for label, ir_map in cast(
        Mapping[str, Mapping[str, float]], res.get("benchmark_ir", {})
    ).items():
        col = f"ir_{label}"
        df[col] = pd.Series(
            {
                k: v
                for k, v in ir_map.items()
                if k not in {"equal_weight", "user_weight"}
            }
        )
    return df


def summary_frame_from_result(res: Mapping[str, object]) -> pd.DataFrame:
    """Return a DataFrame mirroring the Phase-1 summary table."""

    from ..pipeline import _Stats  # lazy import to avoid cycle

    def to_tuple(obj: Any) -> tuple[float, float, float, float, float, float]:
        if isinstance(obj, tuple):
            return cast(tuple[float, float, float, float, float, float], obj)
        s = cast(_Stats, obj)
        return (
            float(s.cagr),
            float(s.vol),
            float(s.sharpe),
            float(s.sortino),
            float(s.information_ratio),
            float(s.max_drawdown),
        )

    def pct(t: Any) -> list[float]:
        a = to_tuple(t)
        return [a[0] * 100, a[1] * 100, a[2], a[3], a[4], a[5] * 100]

    bench_map = cast(Mapping[str, Mapping[str, float]], res.get("benchmark_ir", {}))
    bench_labels = list(bench_map)
    columns = [
        "Name",
        "Weight",
        "IS CAGR",
        "IS Vol",
        "IS Sharpe",
        "IS Sortino",
        "IS IR",
        "IS MaxDD",
        "OS CAGR",
        "OS Vol",
        "OS Sharpe",
        "OS Sortino",
        "OS IR",
    ]
    columns.extend([f"OS IR {b}" for b in bench_labels])
    columns.append("OS MaxDD")
    # Optional trailing AvgCorr if present on in_sample_stats objects (attached in pipeline)
    # Append AvgCorr columns only if stats objects carry correlation values
    try:  # pragma: no cover - data dependent
        in_stats_map = cast(Mapping[str, _Stats], res.get("in_sample_stats", {}))
        if in_stats_map:
            probe = next(iter(in_stats_map.values()))
            # New fields: is_avg_corr / os_avg_corr
            if (
                getattr(probe, "is_avg_corr", None) is not None
                or getattr(probe, "os_avg_corr", None) is not None
            ):
                columns.append("IS AvgCorr")
                columns.append("OS AvgCorr")
    except Exception:  # pragma: no cover - defensive
        pass

    rows: list[list[Any]] = []

    for label, ins, outs in [
        ("Equal Weight", res["in_ew_stats"], res["out_ew_stats"]),
        ("User Weight", res["in_user_stats"], res["out_user_stats"]),
    ]:
        vals = pct(ins) + pct(outs)
        extra = [
            bench_map.get(b, {}).get(
                "equal_weight" if label == "Equal Weight" else "user_weight",
                pd.NA,
            )
            for b in bench_labels
        ]
        rows.append([label, pd.NA, *vals, *extra])

    rows.append([pd.NA] * len(columns))

    include_avg = columns[-1] == "OS AvgCorr"
    for fund, stat_in in cast(Mapping[str, _Stats], res["in_sample_stats"]).items():
        stat_out = cast(Mapping[str, _Stats], res["out_sample_stats"])[fund]
        weight = cast(Mapping[str, float], res["fund_weights"])[fund] * 100
        vals = pct(stat_in) + pct(stat_out)
        extra = [bench_map.get(b, {}).get(fund, pd.NA) for b in bench_labels]
        if include_avg:
            is_ac = getattr(stat_in, "is_avg_corr", None)
            os_ac = getattr(stat_out, "os_avg_corr", None)
            rows.append([fund, weight, *vals, *extra, is_ac, os_ac])
        else:
            rows.append([fund, weight, *vals, *extra])

    return pd.DataFrame(rows, columns=columns)


def combined_summary_result(
    results: Iterable[Mapping[str, object]],
) -> Mapping[str, object]:
    """Return an aggregated result dict across all periods."""

    from collections import defaultdict

    from ..pipeline import _compute_stats, calc_portfolio_returns

    fund_in: dict[str, list[pd.Series]] = defaultdict(list)
    fund_out: dict[str, list[pd.Series]] = defaultdict(list)
    ew_in_series: list[pd.Series] = []
    ew_out_series: list[pd.Series] = []
    user_in_series: list[pd.Series] = []
    user_out_series: list[pd.Series] = []
    weight_sum: dict[str, float] = defaultdict(float)
    periods = 0

    for res in results:
        in_df = cast(pd.DataFrame, res.get("in_sample_scaled"))
        out_df = cast(pd.DataFrame, res.get("out_sample_scaled"))
        ew_map = cast(Mapping[str, float], res.get("ew_weights", {}))
        fund_map = cast(Mapping[str, float], res.get("fund_weights", {}))
        ew_w = [ew_map.get(c, 0.0) for c in in_df.columns]
        user_w = [fund_map.get(c, 0.0) for c in in_df.columns]
        ew_in_series.append(calc_portfolio_returns(np.array(ew_w), in_df))
        ew_out_series.append(calc_portfolio_returns(np.array(ew_w), out_df))
        user_in_series.append(calc_portfolio_returns(np.array(user_w), in_df))
        user_out_series.append(calc_portfolio_returns(np.array(user_w), out_df))
        for c in in_df.columns:
            fund_in[c].append(in_df[c])
            weight_sum[c] += fund_map.get(c, 0.0)
        for c in out_df.columns:
            fund_out[c].append(out_df[c])
    periods += 1

    rf_in = pd.Series(0.0, index=pd.concat(ew_in_series).index)
    rf_out = pd.Series(0.0, index=pd.concat(ew_out_series).index)
    in_ew_stats = _compute_stats(pd.DataFrame({"ew": pd.concat(ew_in_series)}), rf_in)[
        "ew"
    ]
    out_ew_stats = _compute_stats(
        pd.DataFrame({"ew": pd.concat(ew_out_series)}), rf_out
    )["ew"]
    in_user_stats = _compute_stats(
        pd.DataFrame({"user": pd.concat(user_in_series)}), rf_in
    )["user"]
    out_user_stats = _compute_stats(
        pd.DataFrame({"user": pd.concat(user_out_series)}), rf_out
    )["user"]

    # Compute per-fund stats with risk-free series aligned to each fund's
    # concatenated return index to avoid shape mismatches when a fund is
    # not present in every period.
    # Use a broad type to avoid import cycles and undefined-name issues during linting.
    in_stats: dict[str, Any] = {}
    for f, series_list in fund_in.items():
        joined = pd.concat(series_list)
        rf = pd.Series(0.0, index=joined.index)
        in_stats[f] = _compute_stats(pd.DataFrame({f: joined}), rf)[f]

    out_stats: dict[str, Any] = {}
    for f, series_list in fund_out.items():
        joined = pd.concat(series_list)
        rf = pd.Series(0.0, index=joined.index)
        out_stats[f] = _compute_stats(pd.DataFrame({f: joined}), rf)[f]

    fund_weights = {f: weight_sum[f] / periods for f in weight_sum}

    out: dict[str, Any] = {
        "in_ew_stats": in_ew_stats,
        "out_ew_stats": out_ew_stats,
        "in_user_stats": in_user_stats,
        "out_user_stats": out_user_stats,
        "in_sample_stats": in_stats,
        "out_sample_stats": out_stats,
        "fund_weights": fund_weights,
        "benchmark_ir": {},
    }
    return out


def manager_contrib_table(
    results: Iterable[Mapping[str, object]],
) -> pd.DataFrame:
    """Compute per-manager participation and contribution across periods.

    Returns a DataFrame with columns:
      - Manager: fund/manager name
      - Years: total years in portfolio (months with positive weight / 12)
      - OOS CAGR: annualized return of the manager while held
      - Contribution Share: share of total portfolio return contributed
    """

    from collections import defaultdict

    months_held: dict[str, int] = defaultdict(int)
    series_map: dict[str, list[pd.Series]] = defaultdict(list)
    contrib_sum: dict[str, float] = defaultdict(float)
    total_contrib = 0.0

    for res in results:
        out_df = cast(pd.DataFrame | None, res.get("out_sample_scaled"))
        if out_df is None or out_df.empty:
            continue
        weights = cast(Mapping[str, float], res.get("fund_weights", {}))
        # Consider only managers present with positive weight in this period
        for fund in out_df.columns:
            w = float(weights.get(fund, 0.0))
            if w <= 0.0:
                continue
            s = out_df[fund].dropna()
            if s.empty:
                continue
            months_held[fund] += int(s.shape[0])
            series_map[fund].append(s)
            c = float((s * w).sum())
            contrib_sum[fund] += c
            total_contrib += c

    rows: list[dict[str, Any]] = []
    for fund, months in months_held.items():
        concat = (
            pd.concat(series_map[fund]) if series_map[fund] else pd.Series(dtype=float)
        )
        n = int(concat.shape[0])
        if n > 0:
            s_float = concat.astype(float)
            gross = float(np.prod(1.0 + s_float.to_numpy(dtype=float)))
            cagr = float(gross ** (12.0 / n) - 1.0)
        else:
            cagr = float("nan")
        share = (contrib_sum[fund] / total_contrib) if total_contrib != 0.0 else 0.0
        rows.append(
            {
                "Manager": fund,
                "Years": months / 12.0,
                "OOS CAGR": cagr,
                "Contribution Share": share,
            }
        )

    if not rows:
        return pd.DataFrame(
            columns=["Manager", "Years", "OOS CAGR", "Contribution Share"]
        )

    df = pd.DataFrame(rows)
    # Order by contribution share descending, then by manager name
    df.sort_values(
        ["Contribution Share", "Manager"], ascending=[False, True], inplace=True
    )
    df.reset_index(drop=True, inplace=True)
    return df


def combined_summary_frame(results: Iterable[Mapping[str, object]]) -> pd.DataFrame:
    """Return the summary frame across all ``results``."""

    summary = combined_summary_result(results)
    return summary_frame_from_result(summary)


def period_frames_from_results(
    results: Iterable[Mapping[str, object]],
) -> OrderedDict[str, pd.DataFrame]:
    """Return a mapping of sheet names to summary frames for each period."""

    frames: OrderedDict[str, pd.DataFrame] = OrderedDict()
    for idx, res in enumerate(results, start=1):
        period = res.get("period")
        if isinstance(period, (list, tuple)) and len(period) >= 4:
            sheet = str(period[3])
        else:
            sheet = f"period_{idx}"
        frames[sheet] = summary_frame_from_result(res)
    return frames


def workbook_frames_from_results(
    results: Iterable[Mapping[str, object]],
) -> OrderedDict[str, pd.DataFrame]:
    """Return per-period frames plus summary and execution metrics."""

    results_list = list(results)
    frames = period_frames_from_results(results_list)
    if results_list:
        summary = combined_summary_result(results_list)
        frames["summary"] = summary_frame_from_result(summary)
        frames["execution_metrics"] = execution_metrics_frame(results_list)
    return frames


def phase1_workbook_data(
    results: Iterable[Mapping[str, object]],
    *,
    include_metrics: bool = False,
) -> OrderedDict[str, pd.DataFrame]:
    """Return sheet data for a Phase‑1 style multi‑period workbook.

    Parameters
    ----------
    results:
        Iterable of result dictionaries as produced by
        :func:`multi_period.engine.run`.
    include_metrics:
        If ``True`` also return the raw metrics for each period and the
        combined summary.

    Returns
    -------
    OrderedDict[str, pd.DataFrame]
        Mapping of sheet names to data frames suitable for
        :func:`export_to_excel`.
    """

    results_list = list(results)
    frames = workbook_frames_from_results(results_list)

    if include_metrics:
        metrics_frames: OrderedDict[str, pd.DataFrame] = OrderedDict()
        for idx, res in enumerate(results_list, start=1):
            period = res.get("period")
            sheet = (
                str(period[3])
                if isinstance(period, (list, tuple)) and len(period) >= 4
                else f"period_{idx}"
            )
            metrics_frames[f"metrics_{sheet}"] = metrics_from_result(res)

        if results_list and "summary" in frames:
            summary = combined_summary_result(results_list)
            metrics_frames["metrics_summary"] = metrics_from_result(summary)

        frames.update(metrics_frames)

    return frames


def flat_frames_from_results(
    results: Iterable[Mapping[str, object]],
) -> dict[str, pd.DataFrame]:
    """Return consolidated period and summary frames for CSV/JSON export."""

    results_list = list(results)
    frames = workbook_frames_from_results(results_list)
    period_frames = [
        (k, v) for k, v in frames.items() if k not in {"summary", "execution_metrics"}
    ]
    combined_frames = []
    for name, df in period_frames:
        df = df.copy()
        df.insert(0, "Period", name)
        combined_frames.append(df)
    combined = (
        pd.concat(combined_frames, ignore_index=True)
        if combined_frames
        else pd.DataFrame()
    )
    out: dict[str, pd.DataFrame] = {"periods": combined}
    if "summary" in frames:
        out["summary"] = frames["summary"]
        contrib_df = manager_contrib_table(results_list)
        if not contrib_df.empty:
            out["manager_contrib"] = contrib_df
    out["execution_metrics"] = execution_metrics_frame(results_list)

    # Also emit a combined manager changes frame if available
    changes_rows: list[dict[str, Any]] = []
    for res in results_list:
        period = res.get("period")
        period_label = (
            str(period[3])
            if isinstance(period, (list, tuple)) and len(period) >= 4
            else ""
        )
        for ev in (
            cast(list[Mapping[str, Any]] | None, res.get("manager_changes")) or []
        ):
            row: dict[str, Any] = {"Period": period_label}
            for k in ["action", "manager", "firm", "reason", "detail"]:
                val = ev.get(k) if isinstance(ev, Mapping) else None
                row[k] = val if val is not None else ""
            changes_rows.append(row)
    if changes_rows:
        out["changes"] = pd.DataFrame(changes_rows)
    return out


def export_phase1_workbook(
    results: Iterable[Mapping[str, object]],
    output_path: str,
    *,
    include_metrics: bool = False,
) -> None:
    """Export a Phase-1 style workbook for ``results``.

    Each period becomes its own sheet and a final ``summary`` sheet aggregates
    portfolio returns across all periods using the same formatting.
    """

    results_list = list(results)
    reset_formatters_excel()
    frames = phase1_workbook_data(results_list, include_metrics=include_metrics)

    # Register the period sheet formatters
    for idx, res in enumerate(results_list, start=1):
        period = res.get("period")
        if isinstance(period, (list, tuple)) and len(period) >= 4:
            in_s, in_e, out_s, out_e = map(str, period[:4])
            sheet = str(period[3])
        else:
            in_s = in_e = out_s = out_e = ""
            sheet = f"period_{idx}"
        make_period_formatter(sheet, res, in_s, in_e, out_s, out_e)

    # Register the summary formatter if applicable
    if results_list and "summary" in frames:
        summary = combined_summary_result(results_list)
        contrib_df = manager_contrib_table(results_list)
        first = results_list[0].get("period")
        last = results_list[-1].get("period")
        if isinstance(first, (list, tuple)) and isinstance(last, (list, tuple)):
            # Aggregate manager changes across periods with Period labels
            changes_rows: list[dict[str, Any]] = []
            for res in results_list:
                period = res.get("period")
                period_label = (
                    str(period[3])
                    if isinstance(period, (list, tuple)) and len(period) >= 4
                    else ""
                )
                for ev in (
                    cast(list[Mapping[str, Any]] | None, res.get("manager_changes"))
                    or []
                ):
                    row: dict[str, Any] = {"Period": period_label}
                    for k in ["action", "manager", "firm", "reason", "detail"]:
                        val = ev.get(k) if isinstance(ev, Mapping) else None
                        row[k] = val if val is not None else ""
                    changes_rows.append(row)
            summary_ext = dict(summary)
            if changes_rows:
                summary_ext["manager_changes"] = changes_rows
            if not contrib_df.empty:
                summary_ext["manager_contrib"] = contrib_df
            make_summary_formatter(
                summary_ext,
                str(first[0]),
                str(first[1]),
                str(last[2]),
                str(last[3]),
            )
        else:
            summary_ext = dict(summary)
            if not contrib_df.empty:
                summary_ext["manager_contrib"] = contrib_df
            make_summary_formatter(summary_ext, "", "", "", "")
    # Reorder sheets to place summary first if present
    if "summary" in frames:
        ordered: OrderedDict[str, pd.DataFrame] = OrderedDict()
        ordered["summary"] = frames["summary"]
        if "execution_metrics" in frames:
            ordered["execution_metrics"] = frames["execution_metrics"]
        for k, v in frames.items():
            if k in {"summary", "execution_metrics"}:
                continue
            ordered[k] = v
        frames = ordered

    export_to_excel(frames, output_path)


def export_phase1_multi_metrics(
    results: Iterable[Mapping[str, object]],
    output_path: str,
    *,
    formats: Iterable[str] = ("xlsx",),
    include_metrics: bool = False,
) -> None:
    """Export Phase-1 style metrics for multiple periods.

    Excel workbooks contain one sheet per period plus a ``summary`` sheet.
    CSV and JSON outputs consolidate all period tables into a single
    ``*_periods.*`` file alongside a ``*_summary.*`` file.
    """

    results_list = list(results)
    excel_formats = [f for f in formats if f.lower() in {"excel", "xlsx"}]
    other_formats = [f for f in formats if f.lower() not in {"excel", "xlsx"}]

    if excel_formats:
        path = str(Path(output_path).with_suffix(".xlsx"))
        export_phase1_workbook(results_list, path, include_metrics=include_metrics)

    if other_formats:
        other_data = flat_frames_from_results(results_list)
        if "summary" in other_data and include_metrics:
            other_data["metrics_summary"] = metrics_from_result(
                combined_summary_result(results_list)
            )
        if include_metrics:
            metrics_frames: list[pd.DataFrame] = []
            for idx, res in enumerate(results_list, start=1):
                period = res.get("period")
                sheet = (
                    str(period[3])
                    if isinstance(period, (list, tuple)) and len(period) >= 4
                    else f"period_{idx}"
                )
                metrics = metrics_from_result(res)
                metrics.insert(0, "Period", sheet)
                metrics_frames.append(metrics)
            if metrics_frames:
                other_data["metrics"] = pd.concat(metrics_frames, ignore_index=True)

        export_data(other_data, output_path, formats=other_formats)


def export_multi_period_metrics(
    results: Iterable[Mapping[str, object]],
    output_path: str,
    *,
    formats: Iterable[str] = ("xlsx",),
    include_metrics: bool = False,
) -> None:
    """Export per-period metrics using the canonical exporters.

    Parameters
    ----------
    results:
        Sequence of result dictionaries as produced by
        :func:`multi_period.engine.run`.
    output_path:
        File path prefix for the exported artefacts.
    formats:
        Output formats understood by :func:`export_data`.
    include_metrics:
        If ``True`` also emit the raw metrics frame for each period,
        mirroring the single-period "metrics" sheet.
    """

    excel_formats = [f for f in formats if f.lower() in {"excel", "xlsx"}]
    other_formats = [f for f in formats if f.lower() not in {"excel", "xlsx"}]
    excel_data: dict[str, pd.DataFrame] = {}
    other_data: dict[str, pd.DataFrame] = {}
    reset_formatters_excel()

    results_list = list(results)
    frames: dict[str, pd.DataFrame] = (
        dict(workbook_frames_from_results(results_list)) if results_list else {}
    )

    if other_formats:
        other_data.update(flat_frames_from_results(results_list))
        if include_metrics:
            metrics_frames: list[pd.DataFrame] = []
            for idx, res in enumerate(results_list, start=1):
                period = res.get("period")
                sheet = (
                    str(period[3])
                    if isinstance(period, (list, tuple)) and len(period) >= 4
                    else f"period_{idx}"
                )
                metrics = metrics_from_result(res)
                metrics.insert(0, "Period", sheet)
                metrics_frames.append(metrics)
            if metrics_frames:
                other_data["metrics"] = pd.concat(metrics_frames, ignore_index=True)
            if results_list:
                other_data["metrics_summary"] = metrics_from_result(
                    combined_summary_result(results_list)
                )

    if excel_formats:
        excel_data.update({k: v for k, v in frames.items()})

        for idx, res in enumerate(results_list, start=1):
            period = res.get("period")
            if isinstance(period, (list, tuple)) and len(period) >= 4:
                in_s, in_e, out_s, out_e = map(str, period[:4])
                sheet = str(period[3])
            else:
                in_s = in_e = out_s = out_e = ""
                sheet = f"period_{idx}"
            make_period_formatter(sheet, res, in_s, in_e, out_s, out_e)
            if include_metrics:
                excel_data[f"metrics_{sheet}"] = metrics_from_result(res)

        if results_list and "summary" in frames:
            summary = combined_summary_result(results_list)
            contrib_df = manager_contrib_table(results_list)
            first = results_list[0].get("period")
            last = results_list[-1].get("period")
            if isinstance(first, (list, tuple)) and isinstance(last, (list, tuple)):
                summary_ext = dict(summary)
                if not contrib_df.empty:
                    summary_ext["manager_contrib"] = contrib_df
                make_summary_formatter(
                    summary_ext,
                    str(first[0]),
                    str(first[1]),
                    str(last[2]),
                    str(last[3]),
                )
            else:
                summary_ext = dict(summary)
                if not contrib_df.empty:
                    summary_ext["manager_contrib"] = contrib_df
                make_summary_formatter(summary_ext, "", "", "", "")
            if include_metrics:
                excel_data["metrics_summary"] = metrics_from_result(summary)
        # Reorder sheets to place summary first if present
        if "summary" in excel_data:
            ordered: OrderedDict[str, pd.DataFrame] = OrderedDict()
            ordered["summary"] = excel_data["summary"]
            for k, v in excel_data.items():
                if k != "summary":
                    ordered[k] = v
            excel_data = ordered

        export_data(excel_data, output_path, formats=excel_formats)
    if other_formats:
        export_data(other_data, output_path, formats=other_formats)


def export_data(
    data: Mapping[str, pd.DataFrame],
    output_path: str,
    formats: Iterable[str],
    formatter: Formatter | None = None,
) -> None:
    """Export ``data`` to the specified ``formats``."""
    for fmt in formats:
        fmt_norm = fmt.lower()
        fmt_norm = "xlsx" if fmt_norm == "excel" else fmt_norm
        exporter = EXPORTERS.get(fmt_norm)
        if exporter is None:
            raise ValueError(f"Unsupported format: {fmt}")
        path = str(Path(output_path).with_suffix(f".{fmt_norm}"))
        exporter(data, path, formatter)


__all__ = [
    "FORMATTERS_EXCEL",
    "register_formatter_excel",
    "reset_formatters_excel",
    "make_summary_formatter",
    "make_period_formatter",
    "format_summary_text",
    "export_to_excel",
    "export_to_csv",
    "export_to_json",
    "export_to_txt",
    "export_data",
    "metrics_from_result",
    "combined_summary_result",
    "combined_summary_frame",
    "summary_frame_from_result",
    "period_frames_from_results",
    "workbook_frames_from_results",
    "phase1_workbook_data",
    "flat_frames_from_results",
    "export_phase1_workbook",
    "export_phase1_multi_metrics",
    "export_multi_period_metrics",
    "export_bundle",
    "bundle",
]
